# -*- coding: UTF-8 -*-

# Created by Dr. Jan Philipp Menzel 
# Program: Queue multiple analysis of TG-FA quantities including TG-FA sum profiles
# First version created 2023 11 09
# Notes: Definition of replicate naming pattern both here and in actual XIC analysis script
# Notes: This program starts instances of batch files, which run Skyline runner analyses to generate XIC tsv files as input for the XIC analysis script SKYLITE_8
import math
import datetime
import pandas as pd
import scipy
from scipy import stats
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import subprocess
import statistics

print('This program carries out the analysis of TG-FA fragments of triglycerides as part of the Skylite workflow.')
print('The results of the positive mode analysis (Skylite 5 and 6) and raw data are required to perform this calculation.')
print('Before running, edit or add the samplelist and set the parameter accordingly, line 25.')
#print('Initially, set analyseall=1, to analyse all samples (edit samplelist / replist prior to starting). Afterwards, set to 0 to generate final report files.')

samplelist=1

beforeall=datetime.datetime.now()
if samplelist==0:
	replist=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
			'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_', 
			'FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
			'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
			'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
			'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
			'1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
			'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_',]
elif samplelist==1:
	replist=['FB1_', 'FB2_', 'FB3_', 'NIST_QC1_']

analyseall=1			# switch to either run full analysis or skip and proceed to merge existing results
if analyseall==1:
	repi=0
	while repi<len(replist):
		# send replicate to batch file to start respective Skyline runner analysis and subsequent XIC analysis
		replicate=str(replist[repi])+str(repi)
		subprocess.call([r'.\TGFA_XIC.bat', replicate])
		repi=repi+1

	print('The analysis is completed.')
	analyseall=0
if analyseall==0:
	# begin combine results in two new sheets, one for TG-FA quantities and one for sum FA profile quantities
	after=datetime.datetime.now()
	after=str(after)
	today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'
	infilename=today+'SKYLITE_8_TG_FA_quantities.xlsx'

	tgfasuperlist=[]
	fasuperlist=[]
	toprowsuper=[]
	toprowsuper.append('FA')
	elist=[]
	elist.append('')
	cni=0
	while cni<len(replist):
		toprowsuper.append(replist[cni][:-1])
		elist.append('')
		cni=cni+1
	tgfasuperlist.append(toprowsuper)
	fasuperlist.append(toprowsuper)

	alltgfalist=[]	# list of all TGFA species in superlist
	allfalist=[]	# list of all FA species in superlist

	def appendtgfa(repi, ctgfaname, ctgfaq):
		celist=[]		# define list from scratch to avoid overwriting problems
		celist.append('')
		cni=0
		while cni<len(replist):
			celist.append('')
			cni=cni+1
		tgfasuperlist.append(celist)
		tgfasuperlist[len(tgfasuperlist)-1][0]=ctgfaname
		tgfasuperlist[len(tgfasuperlist)-1][repi+1]=ctgfaq
		#alltgfalist.append(ctgfaname)
		return(tgfasuperlist)
	
	def inserttgfa(repi, ctgfaname, ctgfaq, csi):
		celist=[]		# define list from scratch to avoid overwriting problems
		celist.append('')
		cni=0
		while cni<len(replist):
			celist.append('')
			cni=cni+1
		tgfasuperlist.insert(csi, celist)
		tgfasuperlist[csi][0]=ctgfaname
		tgfasuperlist[csi][repi+1]=ctgfaq
		#alltgfalist.append(ctgfaname)
		return(tgfasuperlist)
	
	def appendfa(repi, cfaname, cfaq):
		celist=[]		# define list from scratch to avoid overwriting problems
		celist.append('')
		cni=0
		while cni<len(replist):
			celist.append('')
			cni=cni+1
		fasuperlist.append(celist)
		fasuperlist[len(fasuperlist)-1][0]=cfaname
		fasuperlist[len(fasuperlist)-1][repi+1]=cfaq
		#alltgfalist.append(ctgfaname)
		return(fasuperlist)
	
	def insertfa(repi, cfaname, cfaq, csi):
		celist=[]		# define list from scratch to avoid overwriting problems
		celist.append('')
		cni=0
		while cni<len(replist):
			celist.append('')
			cni=cni+1
		fasuperlist.insert(csi, celist)
		fasuperlist[csi][0]=cfaname
		fasuperlist[csi][repi+1]=cfaq
		#alltgfalist.append(ctgfaname)
		return(fasuperlist)


	wbtgfa=openpyxl.load_workbook(infilename)		# read output file
	repi=0
	while repi<len(replist):
		#print('         ###########################           ##############')
		#print(tgfasuperlist)
		crepname=replist[repi]
		#print(crepname)

		csheetname='TG_FA_quantities_'+crepname[:-1]
		csheetnamepr='TG_FA_profile_'+crepname[:-1]
		#print(csheetname)
		#print(csheetnamepr)
		cwstgfa=wbtgfa[csheetname]
		cwsfa=wbtgfa[csheetnamepr]
		#add relevant data to superlists
		# begin go through current sample results, first through TGFA to extract data
		rw=2
		gorow=1
		while gorow==1:
			ctgfaname=cwstgfa.cell(row=rw, column=1).value
			#print(ctgfaname)
			if ctgfaname is None:
				gorow=0
			elif ctgfaname=='':
				gorow=0
			elif str(ctgfaname)=='None':
				gorow=0
			elif str(ctgfaname)=='NaN':
				gorow=0
			else:
				ctgfaq=cwstgfa.cell(row=rw, column=4).value
				#print(alltgfalist)
				if str(ctgfaname) in alltgfalist:
					#find index where to add quantity for this sample in superlist
					csi=1
					while csi<len(tgfasuperlist):
						if str(tgfasuperlist[csi][0])==str(ctgfaname):
							cfai=csi
							tgfasuperlist[cfai][repi+1]=ctgfaq
							csi=len(tgfasuperlist)
						csi=csi+1
					#print(tgfasuperlist)
					#print(ctgfaname)
					#print(rw)
					#print('edit code')
					#quit()
				else:
					if repi==0:
						#call appendtgfa function with arguments repi, tgfaname, ctgfaq
						tgfasuperlist=appendtgfa(repi, ctgfaname, ctgfaq)
						alltgfalist.append(ctgfaname)
						#print(tgfasuperlist)
					else:
						# find where to insert new tgfa species list
						csi=1
						while csi<len(tgfasuperlist):
							if csi==(len(tgfasuperlist)-1):		# append as last element, if correct position not determined in loop
								tgfasuperlist=appendtgfa(repi, ctgfaname, ctgfaq)
								alltgfalist.append(ctgfaname)
							if (int(tgfasuperlist[csi][0][3])*10)+(int(tgfasuperlist[csi][0][4]))>(int(ctgfaname[3])*10)+(int(ctgfaname[4])):
								tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csi)
								alltgfalist.append(ctgfaname)
								csi=len(tgfasuperlist)
							elif (int(tgfasuperlist[csi][0][3])*10)+(int(tgfasuperlist[csi][0][4]))==(int(ctgfaname[3])*10)+(int(ctgfaname[4])):
								if str(ctgfaname[8])=='_':
									ctgdb=(int(ctgfaname[6])*10)+int(ctgfaname[7])
								else:
									ctgdb=int(ctgfaname[6])
								if str(tgfasuperlist[csi][0][8])=='_':
									ictgdb=(int(tgfasuperlist[csi][0][6])*10)+int(tgfasuperlist[csi][0][7])
								else:
									ictgdb=int(tgfasuperlist[csi][0][6])

								if ictgdb>ctgdb:	# equal #C and #DB to insert smaller than smallest present DB
									tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csi)
									alltgfalist.append(ctgfaname)
									csi=len(tgfasuperlist)
								else:
									# search for correct place to insert by DB and FA
									csj=csi
									gos=1
									while gos==1:
										if str(ctgfaname[8])=='_':
											ctgdb=(int(ctgfaname[6])*10)+int(ctgfaname[7])
										else:
											ctgdb=int(ctgfaname[6])
										if str(tgfasuperlist[csj][0][8])=='_':
											ictgdb=(int(tgfasuperlist[csj][0][6])*10)+int(tgfasuperlist[csj][0][7])
										else:
											ictgdb=int(tgfasuperlist[csj][0][6])

										if csj==(len(tgfasuperlist)-1):
											tgfasuperlist=appendtgfa(repi, ctgfaname, ctgfaq)
											alltgfalist.append(ctgfaname)
											csj=len(tgfasuperlist)
											csi=len(tgfasuperlist)
											gos=0
										elif (int(tgfasuperlist[csj][0][3])*10)+(int(tgfasuperlist[csj][0][4]))>(int(ctgfaname[3])*10)+(int(ctgfaname[4])):		
											tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csj)
											# insert as no other with same or higher #DB
											alltgfalist.append(ctgfaname)
											csj=len(tgfasuperlist)
											csi=len(tgfasuperlist)
											gos=0
										elif ictgdb>ctgdb:	# equal #C and #DB to insert smaller than smallest present DB
											tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csj)
											alltgfalist.append(ctgfaname)
											csj=len(tgfasuperlist)
											csi=len(tgfasuperlist)
											gos=0
										else:
											if ictgdb==ctgdb:
												# same sum composition, FA to be determined
												if len(ctgfaname)==15:
													icfac=(int(ctgfaname[10])*10)+int(ctgfaname[11])
													icfadb=int(ctgfaname[13])
												elif len(ctgfaname)==14:
													if str(ctgfaname[8])=='_':
														icfac=int(ctgfaname[10])
														icfadb=int(ctgfaname[12])
													else:
														icfac=(int(ctgfaname[9])*10)+int(ctgfaname[10])
														icfadb=int(ctgfaname[12])
												elif len(ctgfaname)==13:
													icfac=int(ctgfaname[9])
													icfadb=int(ctgfaname[11])

												if len(tgfasuperlist[csj][0])==15:
													cfac=(int(tgfasuperlist[csj][0][10])*10)+int(tgfasuperlist[csj][0][11])
													cfadb=int(tgfasuperlist[csj][0][13])
												elif len(tgfasuperlist[csj][0])==14:
													if str(tgfasuperlist[csj][0][8])=='_':
														cfac=int(tgfasuperlist[csj][0][10])
														cfadb=int(tgfasuperlist[csj][0][12])
													else:
														cfac=(int(tgfasuperlist[csj][0][9])*10)+int(tgfasuperlist[csj][0][10])
														cfadb=int(tgfasuperlist[csj][0][12])
												elif len(tgfasuperlist[csj][0])==13:
													cfac=int(tgfasuperlist[csj][0][9])
													cfadb=int(tgfasuperlist[csj][0][11])
												
												if cfac>icfac:
													tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csj)
													alltgfalist.append(ctgfaname)
													csj=len(tgfasuperlist)
													csi=len(tgfasuperlist)
													gos=0
												elif cfac==icfac:
													if cfadb>icfadb:
														tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csj)
														alltgfalist.append(ctgfaname)
														csj=len(tgfasuperlist)
														csi=len(tgfasuperlist)
														gos=0
												#tgfasuperlist=inserttgfa(repi, ctgfaname, ctgfaq, csj)
												#alltgfalist.append(ctgfaname)
												#csj=len(tgfasuperlist)
												#csi=len(tgfasuperlist)
												#gos=0
										csj=csj+1
							csi=csi+1
			rw=rw+1
		
		# end go through current sample results, first through TGFA to extract data
		# begin go through current sample results, second through FA profile to extract data
		rw=2
		gorow=1
		while gorow==1:
			cfaname=cwsfa.cell(row=rw, column=1).value
			#print(cfaname)
			if cfaname is None:
				gorow=0
			elif cfaname=='':
				gorow=0
			elif str(cfaname)=='None':
				gorow=0
			elif str(cfaname)=='NaN':
				gorow=0
			else:
				cfaq=cwsfa.cell(row=rw, column=2).value
				#print(alltgfalist)
				if str(cfaname) in allfalist:
					#find index where to add quantity for this sample in superlist
					csi=1
					while csi<len(fasuperlist):
						if str(fasuperlist[csi][0])==str(cfaname):
							cfai=csi
							fasuperlist[cfai][repi+1]=cfaq
							csi=len(fasuperlist)
						csi=csi+1
				else:
					if repi==0:
						#call appendtgfa function with arguments repi, tgfaname, ctgfaq
						fasuperlist=appendfa(repi, cfaname, cfaq)
						allfalist.append(cfaname)
						#print(tgfasuperlist)
					else:
						# find where to insert new tgfa species list
						csi=1
						while csi<len(fasuperlist):
							
							if len(cfaname)==7:
								cfac=(int(cfaname[3])*10)+(int(cfaname[4]))
								cfadb=int(cfaname[6])
							elif len(cfaname)==6:
								cfac=int(cfaname[3])
								cfadb=int(cfaname[5])
							if len(fasuperlist[csi][0])==7:
								icfac=(int(fasuperlist[csi][0][3])*10)+(int(fasuperlist[csi][0][4]))
								icfadb=int(fasuperlist[csi][0][6])
							elif len(fasuperlist[csi][0])==6:
								icfac=int(fasuperlist[csi][0][3])
								icfadb=int(fasuperlist[csi][0][5])

							if icfac>cfac:
								fasuperlist=insertfa(repi, cfaname, cfaq, csi)
								allfalist.append(cfaname)
								csi=len(fasuperlist)
							elif icfac==cfac:
								if icfadb>cfadb:	# equal #C and #DB to insert smaller than smallest present DB
									fasuperlist=insertfa(repi, cfaname, cfaq, csi)
									allfalist.append(cfaname)
									csi=len(fasuperlist)
							if csi==(len(fasuperlist)-1):		# append as last element, if correct position not determined in loop
								fasuperlist=appendfa(repi, cfaname, cfaq)
								allfalist.append(cfaname)
								csi=len(fasuperlist)-1
							
							csi=csi+1
			rw=rw+1
		# end go through current sample results, second through FA profile to extract data
		if repi==1:
			repi=1#00

		repi=repi+1

	#print(tgfasuperlist)

	wstgfa=wbtgfa.create_sheet('Overview TGFA')
	wstgfapr=wbtgfa.create_sheet('Overview FA profiles')	
	# begin write superlists and relevant volcano plot data
	
	rwi=0
	cli=0
	while cli<len(tgfasuperlist[0]):
		rwi=0
		while rwi<len(tgfasuperlist):
			wstgfa.cell(row=rwi+1, column=cli+1).value=tgfasuperlist[rwi][cli]
			rwi=rwi+1
		cli=cli+1

	rwi=0
	cli=0
	while cli<len(fasuperlist[0]):
		rwi=0
		while rwi<len(fasuperlist):		
			wstgfapr.cell(row=rwi+1, column=cli+1).value=fasuperlist[rwi][cli]
			rwi=rwi+1
		cli=cli+1
	# begin calculate and write volcano plot data
	# begin TGFA volcano plots
	# indeces for data for volcano plots [CCl4, MC4R-KO, CDDA, STAM] - indeces of column in excel file Quantification_isomers
	vstartctrli=[2, 12, 24, 34]
	vsnum=[5, 6, 5, 6]	# number of replicates per mouse model
	emptycol=2
	tvolc=4
	volci=0
	while volci<tvolc:
		#print('...')
		lplist=[]	# lipid list
		pvlist=[]	# raw p value
		ablist=[]	# mean abundance of lipid in set of replicates with higher abundance
		lgfclist=[]	# log2 fold change list
		# Begin calculate means, standard deviations, fold changes and p-values
		snum=vsnum[volci] #repnumvolcano #6	# sample number for statistics calculation (e.g., comparison of 6 WT with 6 KO samples, these have to be the first 6 (12) samples in list negdataname)
		csnum=snum
		go=1
		r=2
		while go==1:
			csnum=snum
			clipid=wstgfa.cell(row=r, column=1).value
			asum=0
			bsum=0
			alist=[]
			blist=[]
			sni=vstartctrli[volci] #0
			while sni<(snum+vstartctrli[volci]):
				caq=wstgfa.cell(row=r, column=sni).value
				if str(caq)=='':
					caq=0
					csnum=csnum-1
				elif caq is None:
					caq=0
					csnum=csnum-1
				else:
					cbq=wstgfa.cell(row=r, column=sni+snum).value
					if str(cbq)=='':
						cbq=0
						csnum=csnum-1
					elif cbq is None:
						cbq=0
						csnum=csnum-1
					else:
						caq=float(caq)
						asum=asum+caq
						alist.append(caq)
						cbq=float(cbq)
						bsum=bsum+cbq
						blist.append(cbq)
				sni=sni+1
			if csnum==0:
				amean=0
				bmean=0
			else:
				amean=asum/csnum
				bmean=bsum/csnum
			#print('...')
			#print(clipid)
			#print(alist)
			#print(blist)

			vci=23*volci
			wstgfa.cell(row=r, column=emptycol+len(replist)+2+vci).value=clipid
			wstgfa.cell(row=r, column=emptycol+len(replist)+3+vci).value=amean    #write 
			wstgfa.cell(row=r, column=emptycol+len(replist)+5+vci).value=bmean    #write 
			if len(alist)<2:
				astd=''
			else:
				astd=statistics.stdev(alist)
			if len(blist)<2:
				bstd=''
			else:
				bstd=statistics.stdev(blist)
			wstgfa.cell(row=r, column=emptycol+len(replist)+4+vci).value=astd    #write 
			wstgfa.cell(row=r, column=emptycol+len(replist)+6+vci).value=bstd    #write 
			if amean==0:
				fcm=0
			else:
				fcm=bmean/amean
				wstgfa.cell(row=r, column=emptycol+len(replist)+7+vci).value=fcm    #write 
			#print(fcm)
			if fcm==0:
				lgfcm=0
				# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
				lplist.append(clipid)
				ablist.append(0)
				pvlist.append(1)
				# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
			else:
				lgfcm=math.log2(fcm)
				wstgfa.cell(row=r, column=emptycol+len(replist)+9+vci).value=lgfcm    #write 
			lgfclist.append(lgfcm)
			if lgfcm==0:
				ok=1
			else:
				# begin calc P value
				t,p=stats.ttest_ind(alist, blist, equal_var=False)
				ts=str(t)
				ps=str(p)
				if ts=='nan':
					t=str(1.0)
				if ps=='nan':
					p=str(1.0)
				t=float(t)
				p=float(p)
				#print(p)
				#print(t)
				# write P value and ttest statistic in excel output file
				wstgfa.cell(row=r, column=emptycol+len(replist)+8+vci).value=p    #write 
				wstgfa.cell(row=r, column=emptycol+len(replist)+12+vci).value=t    #write 
				clgp=math.log10(p)	# p value for volcano plot as -log10()
				clgp=(-1)*clgp
				sign=0
				if p<0.05:		# determine if fold change and p value significant
					if fcm>2:
						sign=1
					elif fcm<0.5:
						sign=1
				if sign==1:
					wstgfa.cell(row=r, column=emptycol+len(replist)+11+vci).value=clgp    #write significant p values / with significant fold changes
				else:
					if clgp==0:
						ok=1
					else:
						wstgfa.cell(row=r, column=emptycol+len(replist)+10+vci).value=clgp    #write insignificant
				# end calc P value
				# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
				lplist.append(clipid)
				if amean>bmean:
					lmean=amean
				else:
					lmean=bmean
				ablist.append(lmean)
				pvlist.append(p)
				# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down

			r=r+1
			clipid=wstgfa.cell(row=r, column=1).value
			if clipid is None:
				go=0
			elif str(clipid)=='NaN':
				go=0
			elif str(clipid)=='':
				go=0
			else:
				go=1


		wstgfa.cell(row=1, column=emptycol+len(replist)+2+vci).value='Lipid'
		wstgfa.cell(row=1, column=emptycol+len(replist)+3+vci).value='Mean [Control]'
		wstgfa.cell(row=1, column=emptycol+len(replist)+4+vci).value='Std. dev. [Control]'
		wstgfa.cell(row=1, column=emptycol+len(replist)+5+vci).value='Mean [MASH]'
		wstgfa.cell(row=1, column=emptycol+len(replist)+6+vci).value='Std. dev. [MASH]'
		wstgfa.cell(row=1, column=emptycol+len(replist)+7+vci).value='Fold change of mean'
		wstgfa.cell(row=1, column=emptycol+len(replist)+8+vci).value='P value'
		wstgfa.cell(row=1, column=emptycol+len(replist)+9+vci).value='log2 fold change'
		wstgfa.cell(row=1, column=emptycol+len(replist)+10+vci).value='="-log10 p-value (not significant)"'
		wstgfa.cell(row=1, column=emptycol+len(replist)+11+vci).value='="-log10 p-value (significant)"'
		wstgfa.cell(row=1, column=emptycol+len(replist)+12+vci).value='t-test statistic'

		# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
		negdataname=replist
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+13+vci).value='P value (Bonferroni)'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+14+vci).value='P value (ASD)'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+15+vci).value='P value (BH)'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+16+vci).value='="-log10 p-value (Bonferroni)"'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+17+vci).value='="-log10 p-value (ASD)"'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+18+vci).value='="-log10 p-value (BH; insignificant)"'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+19+vci).value='="-log10 p-value (BH; also significant after ASD and BF)"'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+20+vci).value='="-log10 p-value (BH; also significant after ASD, but not BF)"'
		wstgfa.cell(row=1, column=emptycol+len(negdataname)+21+vci).value='="-log10 p-value (BH; significant, but not after ASD or BF)"'
		nmc=len(lplist)

		#print('Lists for BH:')
		# begin create sorted lists by p value for BH
		#print(lplist)
		#print(pvlist)
		psortlplist=[]
		psortpvlist=[]
		dlplist=[]
		dpvlist=[]
		ni=0
		while ni<len(lplist):
			dlplist.append(lplist[ni])
			dpvlist.append(pvlist[ni])
			ni=ni+1
		while len(dlplist)>0:
			cpv=min(dpvlist)
			cdi=dpvlist.index(cpv)
			psortpvlist.append(dpvlist[cdi])
			psortlplist.append(dlplist[cdi])
			del dpvlist[cdi]
			del dlplist[cdi]
		#print(lplist)
		#print(pvlist)
		#print(psortlplist)
		#print(psortpvlist)
		# end create sorted lists by p value for BH
		#print('Lists for ASD:')
		# begin create sorted lists by p value for ASD
		#print(lplist)
		#print(pvlist)
		#print(ablist)
		absortlplist=[]
		absortpvlist=[]
		absortablist=[]
		dlplist=[]
		dpvlist=[]
		dablist=[]
		ni=0
		while ni<len(lplist):
			dlplist.append(lplist[ni])
			dpvlist.append(pvlist[ni])
			dablist.append(ablist[ni])
			ni=ni+1
		while len(dlplist)>0:
			cab=min(dablist)
			cdi=dablist.index(cab)
			absortpvlist.append(dpvlist[cdi])
			absortablist.append(dablist[cdi])
			absortlplist.append(dlplist[cdi])
			del dpvlist[cdi]
			del dlplist[cdi]
			del dablist[cdi]
		#print(lplist)
		#print(pvlist)
		#print(absortlplist)
		#print(absortpvlist)
		#print(absortablist)
		# end create sorted lists by p value for ASD
		apvi=0
		while apvi<len(lplist):
			if pvlist[apvi]==1:
				ok=1
			elif pvlist[apvi] is None:
				ok=1
			else:
				# begin Bonferroni
				bfpv=pvlist[apvi]*nmc
				if bfpv>1:
					bfpv=1
				logbfpv=math.log10(bfpv)	# p value for volcano plot as -log10()
				logbfpv=(-1)*logbfpv
				wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+13+vci).value=bfpv
				wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+16+vci).value=logbfpv
				# end Bonferroni

				# begin Abundance-Step-Down
				kasd=(absortlplist.index(lplist[apvi]))
				asdpv=pvlist[apvi]*(nmc-kasd)
				if asdpv>1:
					asdpv=1
				logasdpv=math.log10(asdpv)	# p value for volcano plot as -log10()
				logasdpv=(-1)*logasdpv
				wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+14+vci).value=asdpv
				wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+17+vci).value=logasdpv
				# end Abundance-Step-Down

				# begin Benjamini-Hochberg
				kbh=(psortlplist.index(lplist[apvi]))+1
				bhpv=pvlist[apvi]*(nmc/kbh)
				if bhpv>1:
					bhpv=1
				logbhpv=math.log10(bhpv)	# p value for volcano plot as -log10()
				logbhpv=(-1)*logbhpv
				wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+15+vci).value=bhpv
				# determine colour code (column), depending on: 
				# 1) Insignificant after BH (18)
				# 2) Significant after BH and significant after BF (19)
				# 3) Significant after BH and significant after ASD, but not after BF (20)
				# 4) Significant after BH, but not after ASD or BF (21)
				if abs(lgfclist[apvi])<1:
					if logbhpv==0:
						ok=1
					else:
						wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+18+vci).value=logbhpv	# 1) Insignificant due to small fold change
				else:
					if bhpv>0.05:
						if logbhpv==0:
							ok=1
						else:
							wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+18+vci).value=logbhpv	# 1) Insignificant after BH
					else:
						if bfpv<0.05:
							wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+19+vci).value=logbhpv	# 2) Significant after BH and significant after BF and ASD
						else:
							if asdpv<0.05:
								wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+20+vci).value=logbhpv	# 3) Significant after BH and significant after ASD, but not after BF
							else:
								wstgfa.cell(row=2+apvi, column=emptycol+len(negdataname)+21+vci).value=logbhpv	# 4) Significant after BH, but not after ASD or BF 
				# end Benjamini-Hochberg

			apvi=apvi+1
		# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down

		volci=volci+1

	# end TGFA volcano plots
	# begin FA volcano plots
	vstartctrli=[2, 12, 24, 34]
	vsnum=[5, 6, 5, 6]	# number of replicates per mouse model
	emptycol=2
	tvolc=4
	volci=0
	while volci<tvolc:
		#print('...')
		lplist=[]	# lipid list
		pvlist=[]	# raw p value
		ablist=[]	# mean abundance of lipid in set of replicates with higher abundance
		lgfclist=[]	# log2 fold change list
		# Begin calculate means, standard deviations, fold changes and p-values
		snum=vsnum[volci] #repnumvolcano #6	# sample number for statistics calculation (e.g., comparison of 6 WT with 6 KO samples, these have to be the first 6 (12) samples in list negdataname)
		csnum=snum
		go=1
		r=2
		while go==1:
			csnum=snum
			clipid=wstgfapr.cell(row=r, column=1).value
			asum=0
			bsum=0
			alist=[]
			blist=[]
			sni=vstartctrli[volci] #0
			while sni<(snum+vstartctrli[volci]):
				caq=wstgfapr.cell(row=r, column=sni).value
				if str(caq)=='':
					caq=0
					csnum=csnum-1
				elif caq is None:
					caq=0
					csnum=csnum-1
				else:
					cbq=wstgfapr.cell(row=r, column=sni+snum).value
					if str(cbq)=='':
						cbq=0
						csnum=csnum-1
					elif cbq is None:
						cbq=0
						csnum=csnum-1
					else:
						caq=float(caq)
						asum=asum+caq
						alist.append(caq)
						cbq=float(cbq)
						bsum=bsum+cbq
						blist.append(cbq)
				sni=sni+1
			if csnum==0:
				amean=0
				bmean=0
			else:
				amean=asum/csnum
				bmean=bsum/csnum
			#print('...')
			#print(clipid)
			#print(alist)
			#print(blist)

			vci=23*volci
			wstgfapr.cell(row=r, column=emptycol+len(replist)+2+vci).value=clipid
			wstgfapr.cell(row=r, column=emptycol+len(replist)+3+vci).value=amean    #write 
			wstgfapr.cell(row=r, column=emptycol+len(replist)+5+vci).value=bmean    #write 
			if len(alist)<2:
				astd=''
			else:
				astd=statistics.stdev(alist)
			if len(blist)<2:
				bstd=''
			else:
				bstd=statistics.stdev(blist)
			wstgfapr.cell(row=r, column=emptycol+len(replist)+4+vci).value=astd    #write 
			wstgfapr.cell(row=r, column=emptycol+len(replist)+6+vci).value=bstd    #write 
			if amean==0:
				fcm=0
			else:
				fcm=bmean/amean
				wstgfapr.cell(row=r, column=emptycol+len(replist)+7+vci).value=fcm    #write 
			#print(fcm)
			if fcm==0:
				lgfcm=0
			else:
				lgfcm=math.log2(fcm)
				wstgfapr.cell(row=r, column=emptycol+len(replist)+9+vci).value=lgfcm    #write 
			lgfclist.append(lgfcm)
			if lgfcm==0:
				ok=1
				# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
				lplist.append(clipid)
				ablist.append(0)
				pvlist.append(1)
				# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
			else:
				# begin calc P value
				t,p=stats.ttest_ind(alist, blist, equal_var=False)
				ts=str(t)
				ps=str(p)
				if ts=='nan':
					t=str(1.0)
				if ps=='nan':
					p=str(1.0)
				t=float(t)
				p=float(p)
				#print(p)
				#print(t)
				# write P value and ttest statistic in excel output file
				wstgfapr.cell(row=r, column=emptycol+len(replist)+8+vci).value=p    #write 
				wstgfapr.cell(row=r, column=emptycol+len(replist)+12+vci).value=t    #write 
				clgp=math.log10(p)	# p value for volcano plot as -log10()
				clgp=(-1)*clgp
				sign=0
				if p<0.05:		# determine if fold change and p value significant
					if fcm>2:
						sign=1
					elif fcm<0.5:
						sign=1
				if sign==1:
					wstgfapr.cell(row=r, column=emptycol+len(replist)+11+vci).value=clgp    #write significant p values / with significant fold changes
				else:
					if clgp==0:
						ok=1
					else:
						wstgfapr.cell(row=r, column=emptycol+len(replist)+10+vci).value=clgp    #write insignificant
				# end calc P value
				# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
				lplist.append(clipid)
				if amean>bmean:
					lmean=amean
				else:
					lmean=bmean
				ablist.append(lmean)
				pvlist.append(p)
				# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down

			r=r+1
			clipid=wstgfapr.cell(row=r, column=1).value
			if clipid is None:
				go=0
			elif str(clipid)=='NaN':
				go=0
			elif str(clipid)=='':
				go=0
			else:
				go=1

		
		wstgfapr.cell(row=1, column=emptycol+len(replist)+2+vci).value='FA (Sum of TG-FA)'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+3+vci).value='Mean [Control]'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+4+vci).value='Std. dev. [Control]'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+5+vci).value='Mean [MASH]'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+6+vci).value='Std. dev. [MASH]'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+7+vci).value='Fold change of mean'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+8+vci).value='P value'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+9+vci).value='log2 fold change'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+10+vci).value='="-log10 p-value (not significant)"'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+11+vci).value='="-log10 p-value (significant)"'
		wstgfapr.cell(row=1, column=emptycol+len(replist)+12+vci).value='t-test statistic'

		# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+13+vci).value='P value (Bonferroni)'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+14+vci).value='P value (ASD)'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+15+vci).value='P value (BH)'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+16+vci).value='="-log10 p-value (Bonferroni)"'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+17+vci).value='="-log10 p-value (ASD)"'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+18+vci).value='="-log10 p-value (BH; insignificant)"'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+19+vci).value='="-log10 p-value (BH; also significant after ASD and BF)"'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+20+vci).value='="-log10 p-value (BH; also significant after ASD, but not BF)"'
		wstgfapr.cell(row=1, column=emptycol+len(negdataname)+21+vci).value='="-log10 p-value (BH; significant, but not after ASD or BF)"'
		nmc=len(lplist)

		#print('Lists for BH:')
		# begin create sorted lists by p value for BH
		#print(lplist)
		#print(pvlist)
		psortlplist=[]
		psortpvlist=[]
		dlplist=[]
		dpvlist=[]
		ni=0
		while ni<len(lplist):
			dlplist.append(lplist[ni])
			dpvlist.append(pvlist[ni])
			ni=ni+1
		while len(dlplist)>0:
			cpv=min(dpvlist)
			cdi=dpvlist.index(cpv)
			psortpvlist.append(dpvlist[cdi])
			psortlplist.append(dlplist[cdi])
			del dpvlist[cdi]
			del dlplist[cdi]
		#print(lplist)
		#print(pvlist)
		#print(psortlplist)
		#print(psortpvlist)
		# end create sorted lists by p value for BH
		#print('Lists for ASD:')
		# begin create sorted lists by p value for ASD
		#print(lplist)
		#print(pvlist)
		#print(ablist)
		absortlplist=[]
		absortpvlist=[]
		absortablist=[]
		dlplist=[]
		dpvlist=[]
		dablist=[]
		ni=0
		while ni<len(lplist):
			dlplist.append(lplist[ni])
			dpvlist.append(pvlist[ni])
			dablist.append(ablist[ni])
			ni=ni+1
		while len(dlplist)>0:
			cab=min(dablist)
			cdi=dablist.index(cab)
			absortpvlist.append(dpvlist[cdi])
			absortablist.append(dablist[cdi])
			absortlplist.append(dlplist[cdi])
			del dpvlist[cdi]
			del dlplist[cdi]
			del dablist[cdi]
		#print(lplist)
		#print(pvlist)
		#print(absortlplist)
		#print(absortpvlist)
		#print(absortablist)
		# end create sorted lists by p value for ASD
		apvi=0
		while apvi<len(lplist):
			if pvlist[apvi]==1:
				ok=1
			elif pvlist[apvi] is None:
				ok=1
			else:
				# begin Bonferroni
				bfpv=pvlist[apvi]*nmc
				if bfpv>1:
					bfpv=1
				logbfpv=math.log10(bfpv)	# p value for volcano plot as -log10()
				logbfpv=(-1)*logbfpv
				wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+13+vci).value=bfpv
				wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+16+vci).value=logbfpv
				# end Bonferroni

				# begin Abundance-Step-Down
				kasd=(absortlplist.index(lplist[apvi]))
				asdpv=pvlist[apvi]*(nmc-kasd)
				if asdpv>1:
					asdpv=1
				logasdpv=math.log10(asdpv)	# p value for volcano plot as -log10()
				logasdpv=(-1)*logasdpv
				wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+14+vci).value=asdpv
				wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+17+vci).value=logasdpv
				# end Abundance-Step-Down

				# begin Benjamini-Hochberg
				kbh=(psortlplist.index(lplist[apvi]))+1
				bhpv=pvlist[apvi]*(nmc/kbh)
				if bhpv>1:
					bhpv=1
				logbhpv=math.log10(bhpv)	# p value for volcano plot as -log10()
				logbhpv=(-1)*logbhpv
				wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+15+vci).value=bhpv
				# determine colour code (column), depending on: 
				# 1) Insignificant after BH (18)
				# 2) Significant after BH and significant after BF (19)
				# 3) Significant after BH and significant after ASD, but not after BF (20)
				# 4) Significant after BH, but not after ASD or BF (21)
				if abs(lgfclist[apvi])<1:
					if logbhpv==0:
						ok=1
					else:
						wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+18+vci).value=logbhpv	# 1) Insignificant due to small fold change
				else:
					if bhpv>0.05:
						if logbhpv==0:
							ok=1
						else:
							wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+18+vci).value=logbhpv	# 1) Insignificant after BH
					else:
						if bfpv<0.05:
							wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+19+vci).value=logbhpv	# 2) Significant after BH and significant after BF and ASD
						else:
							if asdpv<0.05:
								wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+20+vci).value=logbhpv	# 3) Significant after BH and significant after ASD, but not after BF
							else:
								wstgfapr.cell(row=2+apvi, column=emptycol+len(negdataname)+21+vci).value=logbhpv	# 4) Significant after BH, but not after ASD or BF 
				# end Benjamini-Hochberg

			apvi=apvi+1
		# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
		volci=volci+1

	# end FA volcano plots
	# end calculate and write volcano plot data
	# end write superlists and relevant volcano plot data

	outfilename=infilename[:-5]+'_all.xlsx'
	wbtgfa.save(outfilename)
	print('Calculation completed. The output file is saved as %s' % outfilename)

# end combine results in two new sheets, one for TG-FA quantities and one for sum FA profile quantities
