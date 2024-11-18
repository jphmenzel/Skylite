# -*- coding: UTF-8 -*-

# Jan Philipp Menzel
# Reads chromatograms from filtered (curated) skyline file 
# and identifies isomers and generates integrals of all species (incl. isomers) from XICs.
# Generates transition list with explicit RT and RT window to view final results (feature list containing isomers that are chromatographically separated) in Skyline
# and visualizes results from quantification estimation
## Notes: 

import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import subprocess
import statistics
import csv
import sys
maxInt=sys.maxsize
beforeall=datetime.datetime.now()
################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
################
beforeall=datetime.datetime.now()
#replist=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
		#'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_']#, 
		#'FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
		#'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
		#'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
		#'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
		#'1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
		#'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_',]		# not required, modify in SKYLITE_9
insequence=1				# MODIFY TO TOGGLE BETWEEN SINGLE ANALYSIS OR MULTI ANALYSIS
if insequence==0:
	crepname='FB1_'
	print('Current replicate: %s' % crepname)
	repi=0
else:
	crepname=str(sys.argv[1]) 	# from argument
	if crepname=='CML1_':							# MODIFY TO TEST FOR FIRST SAMPLE IN QUEUE
		repi=0
	else:
		repi=1
	#crepname=replist[repi]	#modify to grab replicate name from list defined above and repi as passed argument from python script call
	print('Current replicate: %s' % crepname)
checkup=0
cimax=1500
maxisomerfilter=1
# Begin convert tsv file to csv and get data 
#csv.field_size_limit(sys.maxsize)
while True:
	try:
		csv.field_size_limit(maxInt)
		break
	except OverflowError:
		maxInt=int(maxInt/10)
convertfile=1		# set 0 for troubleshooting (run this python script on its own using csv file), 1 is default value for running workflow through batch file
if convertfile==1:
	print('Begin conversion from tsv to csv...')
	# begin convert tsv file generated from Skyline runner to csv file # BEGIN EXTRACT INTENSITIES
	try:
	    with open(r'Skyl_XIC_Rep_JPM_TG_FA_pos.tsv', 'r', newline='\n') as in_f, \
	         open(r'skyl_xic_report_1_intensities_pos.csv', 'w', newline='\n') as out_f:
	        reader = csv.reader(in_f, delimiter='\t')
	        writer = csv.writer(out_f, delimiter=',')
	        for li in reader:
	            try:
	                writer.writerow([li[0], li[1], li[2], li[3], li[4], li[5], li[6], li[7], li[9]])
	            except IndexError:  # Prevent errors on blank lines.
	                pass
	except IOError as err:
	    print(err)
	print('Converted tsv file generated from Skyline runner to csv file.')
	# begin delete double quotes from generated csv file
	with open('skyl_xic_report_1_intensities_pos.csv', "r+", encoding="utf-8") as csv_file:
	    content = csv_file.read()
	with open('skyl_xic_report_1_intensities_pos.csv', "w+", encoding="utf-8") as csv_file:
	    csv_file.write(content.replace('"', ''))
	# end delete double quotes from generated csv file # END EXTRACT INTENSITIES
	# begin convert float values for intensities to integers to reduce file size
	#tempdf=pd.read_csv('skyl_xic_report_1_intensities.csv', header=None, skiprows=1)
	tempdf=pd.read_csv('skyl_xic_report_1_intensities_pos.csv', header=None, skiprows=1, delimiter=',', names=list(range(cimax)), low_memory=False)
	#tempdf=tempdf[0].str.split('\s\|\s', expand=True)
	#print(tempdf)
	templist=tempdf.values.tolist()
	tcol=0
	trow=0
	while trow<(len(templist)):		# replaces content of first column (FileName) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=2
	trow=0
	while trow<(len(templist)):		# replaces content of third column (Precursorcharge) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=6
	trow=0
	while trow<(len(templist)):		# replaces content of seventh column (IsotopeLabel) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=9
	trow=0
	while trow<(len(templist)):		# converts intensities to integers
		tcol=9
		while tcol<(len(templist[0])):
			if str(templist[trow][tcol])=='nan':
				ok=1
			else:
				templist[trow][tcol]=float(templist[trow][tcol])
				templist[trow][tcol]=round(templist[trow][tcol], 0)
				templist[trow][tcol]=int(templist[trow][tcol])
			tcol=tcol+1
		trow=trow+1
	tempconvdf=pd.DataFrame(templist)
	filename='skyl_xic_report_1_intensities_pos.csv'
	tempconvdf.to_csv(filename, index=False)
	templist=[]
	tempconvdf=pd.DataFrame(templist)
	tempdf=pd.DataFrame(templist)
	# end convert float values for intensities to integers to reduce file size
print('Converted intensities.')
if convertfile==1:
	# begin convert tsv file generated from Skyline runner to csv file # BEGIN EXTRACT TIMES
	try:
	    with open(r'Skyl_XIC_Rep_JPM_TG_FA_pos.tsv', 'r', newline='\n') as in_f, \
	         open(r'skyl_xic_report_1_times_pos.csv', 'w', newline='\n') as out_f:
	        reader = csv.reader(in_f, delimiter='\t')
	        writer = csv.writer(out_f, delimiter=',')
	        for li in reader:
	            try:
	                writer.writerow([li[8]])
	            except IndexError:  # Prevent errors on blank lines.
	                pass
	except IOError as err:
	    print(err)
	# end convert tsv file generated from Skyline runner to csv file
	# begin delete double quotes from generated csv file
	with open('skyl_xic_report_1_times_pos.csv', "r+", encoding="utf-8") as csv_file:
	    content = csv_file.read()
	with open('skyl_xic_report_1_times_pos.csv', "w+", encoding="utf-8") as csv_file:
	    csv_file.write(content.replace('"', ''))
	# end delete double quotes from generated csv file
	xictimesdf=pd.read_csv('skyl_xic_report_1_times_pos.csv', header=None, skiprows=1, delimiter=',', names=list(range(cimax)))
	xictimeslistfromdf=xictimesdf.values.tolist()
	#xictimeslist=xictimeslistfromdf[0]		# contains times of XICs
	xd=0
	while xd<len(xictimeslistfromdf):
		aa=0
		while aa<8:
			xictimeslistfromdf[xd].insert(0, 0)
			aa=aa+1
		xd=xd+1
	# begin save times of XICs in csv file
	xictimesconvdf=pd.DataFrame(xictimeslistfromdf) #.transpose()
	filename='skyl_xic_report_1_times_pos.csv'
	xictimesconvdf.to_csv(filename, index=False)
	# end save times of XICs in csv file
	xictimesdf=pd.DataFrame(templist)
	xictimeslistfromdf=[]
	#print(xictimesdf)
	#print('list:')
	#print(xictimeslist)
	#print(len(xictimeslist))
	# END EXTRACT TIMES
print('Converted retention times.')
# conversion into csv is ok, data are in correct places, use precursor and fragments to detect, if multiple isomers present
# correct for artefacts (linear increases in MS2 level data - false data points)
# integrate over all relevant isomers and prepare summary for these 

# begin go through XICs and correct artefacts (linear increases where there are no actual data points)
trdf=pd.read_csv('skyl_xic_report_1_times_pos.csv')
toprowxicrt=[trdf.columns.values.tolist()]
toprowxicrt=toprowxicrt[0]
#trdf=trdf.transpose()
xicrtlist=trdf.values.tolist()
#print('Number of rows in Skyl_Rep_JPM_OxLiD1_pos_manual.csv: %d' % ki)
ntrdf=pd.read_csv('skyl_xic_report_1_intensities_pos.csv', low_memory=False)
toprowxicint=[trdf.columns.values.tolist()]
toprowxicint=toprowxicint[0]
#ntrdf=ntrdf.transpose()																# NOT TRANSPOSED TO ALIGN WITH FPSXICINTLIST
xicintlist=ntrdf.values.tolist()
#print('Number of rows in Skyl_Rep_JPM_OxLiD1_neg_manual.csv: %d' % nki)
#print(xicrtlist[831][0])
#print(xicintlist[831][0])
#crti=825
#cx=3
#print(abs(abs(xicintlist[crti-2][cx]-xicintlist[crti-1][cx])-abs(xicintlist[crti-1][cx]-xicintlist[crti][cx])))
#print(len(xicintlist[0]))
#  					# FEATURE FOR REMOVAL OF INTERNAL LINEAR ARTEFACTS
caf=1
dtol=0.015	# tolerance level for linearity (artefact detection)	# could also likely be set to 0.05 or even 0.1 if some artefacts not removed, otherwise check data
cx=0
while cx<len(xicintlist):	# go through list of transitions (rows)
	milist=[]	# indeces of datapoints marked to be set to 0
	crti=10
	go=1
	if str(xicintlist[cx][crti+2])=='nan':
		go=0
	else:
		go=1
	while go==1:		# go through current XIC, all values
		if str(xicintlist[cx][crti+3])=='nan':
			go=0
		else:
			go=1
		tst=1
		if (abs(xicintlist[cx][crti-2]-xicintlist[cx][crti-1]))==0:
			tst=0
		elif (abs(xicintlist[cx][crti-1]-xicintlist[cx][crti]))==0:
			tst=0
		elif (abs(xicintlist[cx][crti]-xicintlist[cx][crti+1]))==0:
			tst=0
		if tst==1:
			if (abs(abs(xicintlist[cx][crti-2]-xicintlist[cx][crti-1])-abs(xicintlist[cx][crti-1]-xicintlist[cx][crti])))/(abs(xicintlist[cx][crti-2]-xicintlist[cx][crti-1]))<dtol:
				if (abs(abs(xicintlist[cx][crti-1]-xicintlist[cx][crti])-abs(xicintlist[cx][crti]-xicintlist[cx][crti+1])))/(abs(xicintlist[cx][crti-1]-xicintlist[cx][crti]))<dtol:
					if (abs(abs(xicintlist[cx][crti]-xicintlist[cx][crti+1])-abs(xicintlist[cx][crti+1]-xicintlist[cx][crti+2])))/(abs(xicintlist[cx][crti]-xicintlist[cx][crti+1]))<dtol:
						if (crti-1) in milist:
							ok=1
						else:
							milist.append(crti-1)
						if (crti) in milist:
							ok=1
						else:
							milist.append(crti)
						if (crti+1) in milist:
							ok=1
						else:
							milist.append(crti+1)		
		crti=crti+1
	rtd=8
	go=1
	if str(xicintlist[cx][rtd+2])=='nan':
		go=0
	else:
		go=1
	while go==1:
		if rtd in milist:
			xicintlist[cx][rtd]=0			# set marked values to 0 (if index in milist)
		if str(xicintlist[cx][rtd+3])=='nan':
			go=0
		else:
			go=1
		rtd=rtd+1
	cx=cx+1

xicdf=pd.DataFrame(xicintlist)#.transpose()		#print('Transposed')
#xicdf.columns=toprowxicint
filename='skyl_xic_report_1_intensities_CORRECTED_pos.csv'
xicdf.to_csv(filename, index=False)
# end go through XICs and correct artefacts (linear increases where there are no actual data points)
print('Artefacts from chromatograms removed.')
# begin add 5 zero values after each XIC in xicintlist to enable better peak detection
cx=0
while cx<len(xicintlist):
	ci=0
	while ci<5:
		xicintlist[cx].append(0)
		xicrtlist[cx].append(0)
		ci=ci+1
	cx=cx+1
# end add 5 zero values after each XIC in xicintlist to enable better peak detection
# begin run five point average (multiple point average) of XICs 
fpsxicintlist=[] #xicintlist
#print(len(xicintlist[1]))
#print(xicintlist[1])
cx=0
while cx<len(xicintlist):
	cfpsxicintlist=[]
	ci=0
	while ci<10:
		cfpsxicintlist.append(xicintlist[cx][ci])
		ci=ci+1
	go=1
	while go==1:
		if ci>len(xicintlist[cx])-4:
			go=0
		elif str(xicintlist[cx][ci+3])=='nan':
			go=0
		else:
			go=1
		#print(xicintlist[ci+2][cx])
		#print(cx)
		#print(ci)
		fps=(xicintlist[cx][ci-2]+xicintlist[cx][ci-1]+xicintlist[cx][ci]+xicintlist[cx][ci+1]+xicintlist[cx][ci+2])/5
		tps=(xicintlist[cx][ci-1]+xicintlist[cx][ci]+xicintlist[cx][ci+1])/3
		ps=(fps+2*tps+xicintlist[cx][ci])/4
		cfpsxicintlist.append(ps)
		ci=ci+1
		#cfpsxicintlist.append(xicintlist[ci][cx])
		#cfpsxicintlist.append(xicintlist[ci+1][cx])
	ci=ci+3
	while ci<cimax:
		cfpsxicintlist.append(0)
		ci=ci+1
	fpsxicintlist.append(cfpsxicintlist)
	cx=cx+1
fpsxicdf=pd.DataFrame(fpsxicintlist)#.transpose()		#print('Transposed')
filename='skyl_xic_report_1_intensities_FPS_pos.csv'
fpsxicdf.to_csv(filename, index=False)
# end run five point average of XICs 
print('Smoothing of chromatograms done.')
#print(fpsxicintlist[0][1])
#print(fpsxicintlist[0][3])
# Use RT Start End limits from sum composition Skyline curated report to get values relevant for integration
# begin read Skyline report pos TG sum composition to get RT Start and RT End
print('Reading file Skyline_Report_curated_for_SKYLITE_5_pos_input_8.csv to retrieve RT ranges for TGs.')
ntrdf=pd.read_csv('Skyline_Report_curated_for_SKYLITE_5_pos_input_8.csv', low_memory=False)
toprown=[ntrdf.columns.values.tolist()]
toprown=toprown[0]
ntrdf=ntrdf.transpose()
negreport=ntrdf.values.tolist()
# end read Skyline report pos TG sum composition to get RT Start and RT End
# begin read TG quantities at sum composition to be used to calculate FA amounts associated to TGs
qtgfilename='SKYLITE_6_isomer_quantities_pos_input_8.xlsx'	# OUTPUT FILE FROM PREVIOUS STEP IS NOW INPUT FILE
wbqtg=openpyxl.load_workbook(qtgfilename)
wsqtg=wbqtg['Quantification_isomers']
# end read TG quantities at sum composition to be used to calculate FA amounts associated to TGs
# using crepname as current replicate name
tgfaquantlist=[]
tgfaquantlist.append([])	# Merged name, e.g., TG_48:1_(12:0))    [0]
tgfaquantlist.append([])	# TG sum composition, e.g., TG_48:1		[1]
tgfaquantlist.append([])	# FA name, e.g., TG-FA 12:0				[2]
tgfaquantlist.append([])	# Uncorrected integral					[3]
tgfaquantlist.append([])	# Corrected integral					[4]
tgfaquantlist.append([])	# Quantity [nmol/mg protein]			[5]
tgfaquantlist.append([])	# cfrac									[6]
tgfaquantlist.append([])	# cscquant								[7]
tgfaquantlist.append([])	# RT Start								[8]
tgfaquantlist.append([])	# RT End								[9]
#print(tgfaquantlist)
cx=0
while cx<len(fpsxicintlist):		# go through XIC report, check if species is present within RT range, get integral, calc other metrics
	if 'TG-FA' in str(fpsxicintlist[cx][4]):	 ## don't use precursor XICs
		cxic=fpsxicintlist[cx][9:]
		if max(cxic)==0:
			ok=1	# TG-FA fragment not found
		else:
			# begin get RT Start and RT End
			csctg=str(fpsxicintlist[cx][1])
			checkup2=0
			if checkup2==1:
				print(csctg)
				print(str(fpsxicintlist[cx][4]))
			crtstart=0
			crtend=0
			tgfaquantlist[1].append(csctg)	# TG sum composition, e.g., TG_48:1		[1]
			tgfaquantlist[2].append(str(fpsxicintlist[cx][4]))	# FA name, e.g., TG-FA 12:0				[2]
			if len(fpsxicintlist[cx][4])==10:	
				cmergetg=csctg+'_('+fpsxicintlist[cx][4][6]+fpsxicintlist[cx][4][7]+fpsxicintlist[cx][4][8]+fpsxicintlist[cx][4][9]+')'
			elif len(fpsxicintlist[cx][4])==9:	
				cmergetg=csctg+'_('+fpsxicintlist[cx][4][6]+fpsxicintlist[cx][4][7]+fpsxicintlist[cx][4][8]+')'
			tgfaquantlist[0].append(cmergetg)	#merged TG-FA name
			# go through negreport (Skyline report)
			r=0
			while r<len(negreport[0]):
				if str(negreport[0][r])==csctg:
					if str(negreport[6][r])=='precursor':
						if str(crepname) in str(negreport[20][r]):
							crtstart=float(negreport[18][r])
							crtend=float(negreport[19][r])
							r=len(negreport[0])
				r=r+1
			if checkup==1:
				if crtstart==0:
					print('What?')
					quit()
			# end get RT Start and RT End
			# begin calculate integral within RT range
			# get RT indices from retention times list (xicrtlist)
			# begin extract relevant XIC (selectxic)
			selectxicrt=[]
			selectxicint=[]
			rti=8
			gort=1
			while gort>0:
				if xicrtlist[cx][rti] is None:
					gort=0
				elif str(xicrtlist[cx][rti])=='':
					gort=0
				elif rti>(len(xicrtlist[cx])-3):
					gort=0
				else:
					if abs((crtstart+0.00001)-float(xicrtlist[cx][rti]))<0.03:
						gort=2
					if (crtstart+0.00001)>float(xicrtlist[cx][rti]):
						if (crtstart+0.00001)<float(xicrtlist[cx][rti+1]):
							gort=2
					if gort==2:
						selectxicrt.append(float(xicrtlist[cx][rti]))
						if rti>(len(fpsxicintlist[cx])-1):
							selectxicint.append(0)
							gort=0
						else:
							selectxicint.append(float(fpsxicintlist[cx][rti]))
					if (crtend+0.00001)>float(xicrtlist[cx][rti]):
						if (crtend+0.00001)<float(xicrtlist[cx][rti+1]):
							selectxicrt.append(float(xicrtlist[cx][rti]))
							if rti>(len(fpsxicintlist[cx])-1):
								selectxicint.append(0)
							else:
								selectxicint.append(float(fpsxicintlist[cx][rti]))
							gort=0
				rti=rti+1
			# end extract relevant XIC
			# begin calculate integral from selectxic
			cintraw=0
			rti=0
			while rti<(len(selectxicint)-1):
				if ((selectxicrt[rti+1]-selectxicrt[rti])*((selectxicint[rti]+selectxicint[rti+1])/2)) is None:
					ok=1
				elif float(((selectxicrt[rti+1]-selectxicrt[rti])*((selectxicint[rti]+selectxicint[rti+1])/2)))>0:
					cintraw=cintraw+((selectxicrt[rti+1]-selectxicrt[rti])*((selectxicint[rti]+selectxicint[rti+1])/2))
				rti=rti+1
			tgfaquantlist[3].append(cintraw)	# Uncorrected integral	[3]
			if len(fpsxicintlist[cx][4])==10:
				unsatindex=int(fpsxicintlist[cx][4][9])
			elif len(fpsxicintlist[cx][4])==9:
				unsatindex=int(fpsxicintlist[cx][4][8])
			ccorrlist=[1, 1.203042, 1.2879, 0.3202, 0.274727, 0.26653, 0.258333, 0.258333, 0.258333]	# correction factor for unsaturated FA TG-FA fragment
			cintcorr=cintraw*ccorrlist[unsatindex]
			tgfaquantlist[4].append(cintcorr)	# Corrected integral	[4]
			tgfaquantlist[5].append(0)			# placeholder for final quantity
			tgfaquantlist[6].append(0)			# placeholder for cfrac
			tgfaquantlist[7].append(0)			# placeholder for cscquant
			tgfaquantlist[8].append(crtstart)			#
			tgfaquantlist[9].append(crtend)			#
			# end calculate integral from selectxic
	cx=cx+1
print('Completed processing of TG-FA chromatograms. Now calculating final quantities...')
# begin calculate final quantity from corrected integrals of all TG-FA fractions within the same sum composition 
# 		and from the quantity for the sum composition from results file
#print(tgfaquantlist[0])
#print(tgfaquantlist[1])
#print(tgfaquantlist[2])
#print(tgfaquantlist[3])
#print(tgfaquantlist[4])
qx=0
while qx<len(tgfaquantlist[0]):
	tx=0
	cfasum=0
	while tx<len(tgfaquantlist[0]):
		if tgfaquantlist[1][tx]==tgfaquantlist[1][qx]:
			cfasum=cfasum+tgfaquantlist[4][tx]
		tx=tx+1
	#qx=qx+1
	if cfasum==0:
		cfrac=0
	else:
		cfrac=tgfaquantlist[4][qx]/cfasum
	if checkup==1:
		print(cfrac)
	#get quantity of TG sum composition
	gosc=1
	sci=2
	while gosc==1:
		tgsc=wsqtg.cell(row=sci, column=1).value
		if tgsc is None:
			gosc=0
		elif str(tgsc)=='':
			gosc=0
		else:
			if tgsc==str(tgfaquantlist[1][qx]):
				usetgscq=0
				ccol=2
				gocol=1
				while gocol==1:
					tgscq=wsqtg.cell(row=sci, column=ccol).value
					tgscrep=wsqtg.cell(row=1, column=ccol).value
					if tgscrep is None:
						gocol=0
					elif str(tgscrep)=='':
						gocol=0
					else:
						if crepname in str(tgscrep):
							if tgscq is None:
								tgscq=0
							usetgscq=float(tgscq)
							gocol=0
							gosc=0
					ccol=ccol+1
		sci=sci+1
	cscquant=usetgscq
	cfracquant=cfrac*cscquant
	tgfaquantlist[5][qx]=float(cfracquant)	# final quantity
	tgfaquantlist[6][qx]=float(cfrac)	# fraction
	tgfaquantlist[7][qx]=float(cscquant)	# fraction
	qx=qx+1
# end calculate final quantity
#print(tgfaquantlist)

after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'

if insequence==0:
	wbtgfa = Workbook(write_only=True)
	outfilename=today+'SKYLITE_8_TG_FA_quantities.xlsx'
	wbtgfa.save(outfilename)
	wbtgfa=openpyxl.load_workbook(outfilename)
	csheetname='TG_FA_quantities_'+crepname[:-1]
	csheetnamepr='TG_FA_profile_'+crepname[:-1]
	wstgfa=wbtgfa.create_sheet(csheetname)
	wstgfapr=wbtgfa.create_sheet(csheetnamepr)		# sheet for FA profile (reconstructed, sums of TG-FA quantities across all TGs)
	del wbtgfa['Sheet']
	wstgfa.cell(row=1, column=1).value='TG & TG-FA'
	wstgfa.cell(row=1, column=2).value='TG sum composition'
	wstgfa.cell(row=1, column=3).value='TG-FA'
	wstgfa.cell(row=1, column=4).value='TG-associated FA quantity [nmol / mg protein]'
	#wstgfa.cell(row=1, column=5).value='Uncorrected integral'
	#wstgfa.cell(row=1, column=6).value='Corrected integral'
	#wstgfa.cell(row=1, column=7).value='cfrac'
	#wstgfa.cell(row=1, column=8).value='TG sum composition quantity'
	#wstgfa.cell(row=1, column=9).value='RT Start'
	#wstgfa.cell(row=1, column=10).value='RT End'
	wstgfapr.cell(row=1, column=1).value='FA'
	wstgfapr.cell(row=1, column=2).value='Quantity [nmol / mg protein]'
else:
	outfilename=today+'SKYLITE_8_TG_FA_quantities.xlsx'
	if repi==0:		# In case of first replicate, create output file 
		wbtgfa = Workbook(write_only=True)
		wbtgfa.save(outfilename)
	wbtgfa=openpyxl.load_workbook(outfilename)		# read output file
	csheetname='TG_FA_quantities_'+crepname[:-1]
	csheetnamepr='TG_FA_profile_'+crepname[:-1]
	wstgfa=wbtgfa.create_sheet(csheetname)
	wstgfapr=wbtgfa.create_sheet(csheetnamepr)		# sheet for FA profile (reconstructed, sums of TG-FA quantities across all TGs)
	if repi==0:
		del wbtgfa['Sheet']
	wstgfa.cell(row=1, column=1).value='TG & TG-FA'
	wstgfa.cell(row=1, column=2).value='TG sum composition'
	wstgfa.cell(row=1, column=3).value='TG-FA'
	wstgfa.cell(row=1, column=4).value='TG-associated FA quantity [nmol / mg protein]'
	wstgfapr.cell(row=1, column=1).value='FA'
	wstgfapr.cell(row=1, column=2).value='Quantity [nmol / mg protein]'

# write into file
k=0
kr=k
while k<len(tgfaquantlist[0]):
	if tgfaquantlist[5][k]==0:
		ok=1
	else:
		wstgfa.cell(row=kr+2, column=1).value=str(tgfaquantlist[0][k])
		wstgfa.cell(row=kr+2, column=2).value=str(tgfaquantlist[1][k])
		wstgfa.cell(row=kr+2, column=3).value=str(tgfaquantlist[2][k])
		wstgfa.cell(row=kr+2, column=4).value=tgfaquantlist[5][k]
		#wstgfa.cell(row=k+2, column=5).value=tgfaquantlist[3][k]
		#wstgfa.cell(row=k+2, column=6).value=tgfaquantlist[4][k]
		#wstgfa.cell(row=k+2, column=7).value=tgfaquantlist[6][k]
		#wstgfa.cell(row=k+2, column=8).value=tgfaquantlist[7][k]
		#wstgfa.cell(row=k+2, column=9).value=tgfaquantlist[8][k]
		#wstgfa.cell(row=k+2, column=10).value=tgfaquantlist[9][k]
		kr=kr+1			# Do not write values that are 0
	k=k+1

# begin reconstruct overall TG-FA profile
falist=[]
faqlist=[]
kf=0
while kf<len(tgfaquantlist[2]):
	if str(tgfaquantlist[2][kf][3:]) in falist:
		faqlist[falist.index(str(tgfaquantlist[2][kf][3:]))]=faqlist[falist.index(str(tgfaquantlist[2][kf][3:]))]+float(tgfaquantlist[5][kf])
	else:
		falist.append(str(tgfaquantlist[2][kf][3:]))
		faqlist.append(float(tgfaquantlist[5][kf]))
	kf=kf+1
# begin sort FA profile
print(falist)
print(faqlist)

pfalist=[]
pfaqlist=[]
pfalist.append(falist[0])
pfaqlist.append(faqlist[0])
ks=1
while ks<len(falist):
	#print(pfalist)
	kt=0
	#ktgo=1
	while kt<len(pfalist):
		if falist[ks]==pfalist[kt]:
			ok=1
		else:
			if len(falist[ks])==7:
				cfas=((10*int(falist[ks][3]))+int(falist[ks][4]))
			elif len(falist[ks])==6:
				cfas=(int(falist[ks][3]))
			if len(pfalist[kt])==7:
				cfat=((10*int(pfalist[kt][3]))+int(pfalist[kt][4]))
			elif len(pfalist[kt])==6:
				cfat=(int(pfalist[kt][3]))
			#print(cfas)
			#print(cfat)
			if cfas<(cfat+1):
				if cfas==cfat:
					if len(falist[ks])==7:
						ik=6
					elif len(falist[ks])==6:
						ik=5
					if len(pfalist[kt])==7:
						jk=6
					elif len(pfalist[kt])==6:
						jk=5
					wk=0
					gowk=1
					while gowk==1:
						if (kt-wk)<1:
							pfalist.insert((kt-wk), falist[ks])
							pfaqlist.insert((kt-wk), faqlist[ks])
							kt=len(pfalist)
							gowk=0
						else:
							if len(pfalist[kt-wk])==7:
								jk=6
							elif len(pfalist[kt-wk])==6:
								jk=5
							if (int(falist[ks][ik]))<(int(pfalist[kt-wk][jk])):			#ok
								wk=wk+1
								if len(pfalist[kt-wk])==7:
									cfat2=((10*int(pfalist[kt-wk][3]))+int(pfalist[kt-wk][4]))
								elif len(pfalist[kt-wk])==6:
									cfat2=(int(pfalist[kt-wk][3]))
								if cfas>cfat2:
									pfalist.insert((kt-wk+1), falist[ks])
									pfaqlist.insert((kt-wk+1), faqlist[ks])
									gowk=0
									kt=len(pfalist)
							elif (int(falist[ks][ik]))>(int(pfalist[kt-wk][jk])):
								if (kt-wk+1)==len(pfalist):
									pfalist.append(falist[ks])
									pfaqlist.append(faqlist[ks])
								else:
									# check for species after that also have less DB
									pk=1
									if len(pfalist[kt-wk+pk])==7:
										cfatp=((10*int(pfalist[kt-wk+pk][3]))+int(pfalist[kt-wk+pk][4]))
									elif len(pfalist[kt-wk+pk])==6:
										cfatp=(int(pfalist[kt-wk+pk][3]))
									if cfas==cfatp:
										gopc=1
										while gopc==1:
											if len(pfalist[kt-wk+pk])==7:
												jk=6
											elif len(pfalist[kt-wk+pk])==6:
												jk=5
											if (int(falist[ks][ik]))>(int(pfalist[kt-wk+pk][jk])):
												pk=pk+1
												if (kt-wk+pk)==len(pfalist):
													pfalist.append(falist[ks])
													pfaqlist.append(faqlist[ks])
													gopc=0
													kt=len(pfalist)
												else:
													if len(pfalist[kt-wk+pk])==7:
														cfatp3=((10*int(pfalist[kt-wk+pk][3]))+int(pfalist[kt-wk+pk][4]))
													elif len(pfalist[kt-wk+pk])==6:
														cfatp3=(int(pfalist[kt-wk+pk][3]))
													if cfas==cfatp3:
														ok=1
													else:
														pfalist.insert((kt-wk+pk), falist[ks])
														pfaqlist.insert((kt-wk+pk), faqlist[ks])
														kt=len(pfalist)
														gopc=0
												
											else:
												pfalist.insert((kt-wk+pk), falist[ks])
												pfaqlist.insert((kt-wk+pk), faqlist[ks])
												kt=len(pfalist)
												gopc=0
									else:
										pfalist.insert((kt-wk+1), falist[ks])
										pfaqlist.insert((kt-wk+1), faqlist[ks])
								kt=len(pfalist)
								gowk=0
				else:
					pfalist.insert(kt, falist[ks])
					pfaqlist.insert(kt, faqlist[ks])
					kt=len(pfalist)
			elif kt==(len(pfalist)-1):
				pfalist.append(falist[ks])
				pfaqlist.append(faqlist[ks])
				kt=len(pfalist)
		kt=kt+1
		#if kt<len(pfalist):
		#	ktgo=1
		#else:
		#	ktgo=0
	ks=ks+1
# end sort FA profile

print(pfalist)
print(pfaqlist)
kp=0
kpw=0
while kp<len(pfalist):
	if float(pfaqlist[kp])==0:
		kp=kp+1
	else:
		wstgfapr.cell(row=kpw+2, column=1).value=str(pfalist[kp])
		wstgfapr.cell(row=kpw+2, column=2).value=float(pfaqlist[kp])
		kp=kp+1
		kpw=kpw+1
# end reconstruct overall TG-FA profile

wbtgfa.save(outfilename)
print('Calculation completed. The output file is saved as %s' % outfilename)


