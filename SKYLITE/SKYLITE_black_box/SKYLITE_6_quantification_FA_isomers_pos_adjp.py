# -*- coding: UTF-8 -*-

# Jan Philipp Menzel
# Performs: Reading reports from Skyline adjustment of integrals and duplication for quantification of chromatographically resolved isomers
# Note: reads previous output from SKYLITE 5 to determine how some quantities have to be split based on isomers present 
# First version created on 6/12/2023
## Notes: 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import subprocess
import statistics
import csv
import sys
import scipy
from scipy import stats
maxInt=sys.maxsize
beforeall=datetime.datetime.now()

qcheck=0

excelinput=1	# default is the use of SKYLITE_INPUT_samplelist.xlsx as a method of entering metadata into the workflow
if excelinput==1:
	print('--------------------------------------------------------------------------------------------------------')
	print(' S K Y L I T E  6 ')
	print('This program calculates quantities of lipid isomers from curated Skyline report files of the SKYLITE workflow and from the output file of SKYLITE 5.')
	print('The required input Skyline report file has to be named according to SKYLITE_INPUT_samplelist.xlsx')
	print('All input data requested in SKYLITE_INPUT_samplelist.xlsx is required.')
	print('Please refer to the associated publication / preprint and the github page for further information.')
	print('--------------------------------------------------------------------------------------------------------')
	wbin=openpyxl.load_workbook('SKYLITE_INPUT_sample_list.xlsx')
	wsin1=wbin['Sample_list']
	wsin2=wbin['Ultimate_SPLASH_ONE_exceptions']
	wsin3=wbin['Input_file_names']
	wsin4=wbin['Isomer_integration_exceptions']
	internalstandard=wsin1.cell(row=2, column=7).value
	tissuetype=wsin1.cell(row=2, column=3).value
	vollivhom=wsin1.cell(row=2, column=4).value
	uspvolul=wsin1.cell(row=2, column=8).value
	uspvol=float(uspvolul)*0.001*0.001	# convert from uL to L
	ismvolul=wsin1.cell(row=2, column=9).value
	ismvol=float(ismvolul)*0.001*0.001	# convert from uL to L
	samplevol=10*0.001*0.001	# convert from uL to L
	negdataname=[]
	replicateweights=[]
	repproteinweight=[]
	goread=1
	rrow=2
	while goread==1:
		csample=wsin1.cell(row=rrow, column=2).value
		if csample is None:
			goread=0
		elif str(csample)=='nan':
			goread=0
		else:
			negdataname.append(str(csample))
			ctt=wsin1.cell(row=2, column=3).value
			if ctt==1:
				vollivhom=wsin1.cell(row=rrow, column=4).value
				crepwt=wsin1.cell(row=rrow, column=5).value
				if crepwt is None:
					replicateweights.append(10)
				else:
					replicateweights.append(float(crepwt))
				crepprwt=wsin1.cell(row=rrow, column=6).value
				if crepprwt is None:
					repproteinweight.append(1)
				else:
					repproteinweight.append(float(crepprwt))
			elif ctt==0:
				csamplevol=wsin1.cell(row=rrow, column=4).value
				samplevol=csamplevol*0.001*0.001	# convert from uL to L
				replicateweights.append(10)
				repproteinweight.append(1)
		rrow=rrow+1

	ms1splitlist=[]
	goread=1
	rrow=2
	while goread==1:
		csplit=wsin4.cell(row=rrow, column=2).value
		if csplit is None:
			goread=0
		elif str(csplit)=='nan':
			goread=0
		else:
			ms1splitlist.append(str(csplit))
		rrow=rrow+1

	remergelist=[]
	goread=1
	rrow=2
	while goread==1:
		crm=wsin4.cell(row=rrow, column=3).value
		if crm is None:
			goread=0
		elif str(crm)=='nan':
			goread=0
		else:
			remergelist.append(str(crm))
		rrow=rrow+1

	cfn=wsin3.cell(row=22, column=6).value	#row defines file # reading isomer report from Skyline
	filenameskylrep4=str(cfn)+'.csv'
	ntrdf=pd.read_csv(filenameskylrep4, low_memory=False)
	toprown=[ntrdf.columns.values.tolist()]
	toprown=toprown[0]
	ntrdf=ntrdf.transpose()
	negreport=ntrdf.values.tolist()

	cfn2=wsin3.cell(row=24, column=6).value	#row defines file
	filenameq3s4=str(cfn2)+'.xlsx'	# OUTPUT FILE FROM PREVIOUS STEP IS NOW INPUT FILE
	wb=openpyxl.load_workbook(filenameq3s4)
	ws=wb['Quantification']
	ws2=wb.create_sheet('Quantification_isomers')

	snum=0
	repnumvolcano=0
	if qcheck==1:
		print(negdataname)
		print(filenameskylrep4)
		print(filenameq3s4)
		print(ms1splitlist)
		print(remergelist)
		#quit()

else:
	print('--------------------------------------------------------------------------------------------------------')
	print(' S K Y L I T E  6 ')
	print('This program calculates quantities of lipid isomers from curated Skyline report files of the SKYLITE workflow and from the output file of SKYLITE 5.')
	print('The required input files for this program are:')
	print(' - Skyline_Report_isomers_for_SKYLITE_6_pos.csv')
	print(' - SKYLITE_5_quantities_for_SKYLITE_6_pos.xlsx')
	print('Manually enter or modify species in ms1splitlist, if applicable:')
	print('These are species at low abundance with missing or inconsistent MS2 data, but isomers, which are chromatographically separated.')
	print('Please refer to the associated publication / preprint and the github page for information.')
	print('--------------------------------------------------------------------------------------------------------')
	print('########### C H E C K L I S T ####################')
	print('   1) Set tissue type, line 42')
	print('   4) Set mousemodel, line 43')
	print('   5) Edit elif clause for mousemodel, containing sample names, after line 116')
	print('   6) Edit elif clause for ms1splitlist, containing lipid names for exceptions, after line 162')
	print('   6) Edit elif clause for remergelist, containing lipid names for exceptions, after line 173')
	calculate=eval(input('All edited according to checklist? Y=1, N=0 ___'))
	if calculate==0:
		quit()

	tissuetype=1		# liver tissue = 1;  blood plasma = 0	# MODIFY for sample type
	mousemodel=9 #6 is NIST validation # 4 # 5 is all four mouse models		# model 1 = CCl4_HFD # model 2 = MC4R-KO_HFD # model 3 = CDDA_HFD # 4 = STAM
	# order for 5 is # model 1 = CCl4_HFD # model 2 = MC4R-KO_HFD # model 3 = STAM # 4 = CDDA_HFD
	################################################# MODIFY 
	#negdataname=['NIST1', 'NIST2', 'NIST3']		# naming pattern of replicates# MODIFY 
	#negdataname=['CML1', 'CML2', 'CML3', 'CML4', 'CML5', 'MML1', 'MML2', 'MML3', 'MML4', 'MML5', 'NIST_QC1_i1', 'NIST_QC1_i1_2', 'PBQC_CML', 'PBQC_MML']		# naming pattern of replicates
	repnumvolcano=3
	if mousemodel==1: # CCl4_HFD
		repnumvolcano=5	
		negdataname=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
				'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_', 
				'NIST_QC1_i1', 'NIST_QC1_i1_2'] #, 'PBQC_CML', 'PBQC_MML']
	elif mousemodel==2:	# MC4R-KO_HFD
		repnumvolcano=6
		negdataname=['FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
				'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
				'FB4_', 'FB20_', 'FB5_', 'FB6_', 'FB7_', 'FB8_', 'FB17_', 'FB18_', 'FB19_', 'NIST_QC1_', 'NIST_QC2_']		# naming pattern of replicates
	elif mousemodel==3: # CDDA HFD
		# Choline deficient:
		repnumvolcano=5	
		negdataname=['14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
				'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
				'24b_CDDA_6_', '24c_CDDA_6_', 
				'1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
				'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
				'NIST_QC1_', 'NIST_QC2_'] #, '13_CTRL_1_']		# naming pattern of replicates
	elif mousemodel==4:	# STAM
		# STAM:
		repnumvolcano=6
		negdataname=['1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
				'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
				'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
				'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
				'24b_CDDA_6_', '24c_CDDA_6_', 
				'NIST_QC1_', 'NIST_QC2_'] #, '13_CTRL_1_']		# naming pattern of replicates
	elif mousemodel==5:
		# CCl4 HFD mouse model Inselspital ###########################################
		nd1=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
			'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_', 
			'USP_pos_NIST_QC1_i1']
		# MC4R-KO_HCD mouse model Deborah Stroka ######################################
		nd2=['FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
			'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
			'DML_pos_NIST_QC1_']		# naming pattern of replicates
		# Choline deficient CDDA mouse model ######################################
		# Streptozotocin diabetic STAM mouse model ######################################
		nd3=['1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
			'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
			'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
			'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
			'24b_CDDA_6_', '24c_CDDA_6_', 
			'KML_pos_NIST_QC2_']		# naming pattern of replicates
		# use replicateweights for liver tissue, mass in mg of liver tissue used for homogenization, assuming standard protocol (check below for vol of liver homogenate used)
		negdataname=[]
		ndi=0
		while ndi<len(nd1):
			negdataname.append(nd1[ndi])
			ndi=ndi+1
		ndi=0
		while ndi<len(nd2):
			negdataname.append(nd2[ndi])
			ndi=ndi+1
		ndi=0
		while ndi<len(nd3):
			negdataname.append(nd3[ndi])
			ndi=ndi+1
	elif mousemodel==6:
		negdataname=['NIST1_', 'NIST2_', 'NIST3_']
	elif mousemodel==7:
		negdataname=['JQC_', 'JQC2_', 'JQC3_']
	elif mousemodel==8:			# liver interday validation
		negdataname=['KML_', 'KML24d_', 'KML24e_']
	elif mousemodel==9:	# MC4R-KO_WD
		repnumvolcano=3
		negdataname=['FB1_', 'FB2_', 'FB3_', 'NIST_QC1_']		# naming pattern of replicates


	else:
		print('Edit/add sampleset and naming pattern with additional elif clause!')
		quit()

	################################################# MODIFY 


	# ms1splitlist contains those species, for which double bond or branching isomers are chromatographically separated with no interference from multiple isomer species and with 
	# low abundance overall (missing or inconsistent MS2 data), so that isomer splitting of quantities according to isomers is based on MS1 precursor integrals of the peaks
	if mousemodel==2:		# MC4R-KO
		ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_33:2_(15:0_18:2)', 'PC_34:3_(16:0_18:3)', 'PC_35:1_(17:0_18:1)', 
					'PC_35:2_(17:0_18:2)', 'PC_36:4_(16:0_20:4)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 
					'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)', 
					'PC_36:3_(16:0_20:3)', 'PC_38:3_(18:0_20:3)']
	elif mousemodel>2:		# CDDA and STAM
		ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_32:1_(16:0_16:1)', 'PC_33:0_(15:0_18:0)', 'PC_33:0_(16:0_17:0)', 
					'PC_33:2_(15:0_18:2)', 'PC_34:3_(16:0_18:3)', 'PC_35:1_(17:0_18:1)', 'PC_35:2_(17:0_18:2)', 
					'PC_36:3_(16:0_20:3)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 'PC_38:3_(18:0_20:3)', 
					'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)']
	elif mousemodel==1:		# CCl4
		ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_33:2_(15:0_18:2)', 'PC_34:3_(16:0_18:3)', 'PC_35:1_(17:0_18:1)', 
					'PC_35:2_(17:0_18:2)', 'PC_36:4_(16:0_20:4)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 
					'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)', 
					'PC_36:3_(16:0_20:3)', 'PC_38:3_(18:0_20:3)', 'PE_38:3_(18:0_20:3)']
	if mousemodel==5:
		ms1splitlist=['DG_36:2_(18:1_18:1)', 'DG_38:6_(18:2_20:4)', 'SM_40:2_(22:0_18:2)']
	elif mousemodel==6:
		ms1splitlist=['SM_34:2_(16:0_18:2)', 'SM_36:3_(18:1_18:2)', 'SM_40:2_(22:0_18:2)', 'SM_41:1_(24:0_17:1)', 'SM_41:2_(17:1_24:1)',
				'SM_42:1_(24:0_18:1)', 'SM_42:2_(24:0_18:2)', 'SM_42:3_(24:1_18:2)', 'SM_43:1_(26:0_17:1)', 'SM_43:2_(22:0_21:2)']
	elif mousemodel==7:
		ms1splitlist=['SM_34:2_(16:0_18:2)', 'SM_36:3_(18:1_18:2)', 'SM_40:2_(22:0_18:2)', 'SM_41:1_(24:0_17:1)', 'SM_41:2_(17:1_24:1)',
				'SM_42:1_(24:0_18:1)', 'SM_42:2_(24:0_18:2)', 'SM_42:3_(24:1_18:2)', 'SM_43:1_(26:0_17:1)', 'SM_43:2_(22:0_21:2)']
	elif mousemodel==8:
		ms1splitlist=['SM_34:2_(16:0_18:2)', 'SM_36:3_(18:1_18:2)', 'SM_40:2_(22:0_18:2)', 'SM_41:1_(24:0_17:1)', 'SM_41:2_(17:1_24:1)',
				'SM_42:1_(24:0_18:1)', 'SM_42:2_(24:0_18:2)', 'SM_42:3_(24:1_18:2)', 'SM_43:1_(26:0_17:1)', 'SM_43:2_(22:0_21:2)', 
				'DG_36:2_(18:1_18:1)', 'DG_36:4_(18:2_18:2)', 'DG_38:6_(18:2_20:4)']
	elif mousemodel==9:		# MC4R-KO
		ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_33:2_(15:0_18:2)', 'PC_34:3_(16:0_18:3)', 'PC_35:1_(17:0_18:1)', 
					'PC_35:2_(17:0_18:2)', 'PC_36:4_(16:0_20:4)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 
					'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)', 
					'PC_36:3_(16:0_20:3)', 'PC_38:3_(18:0_20:3)']
	# remergelist contains those species that have faulty quantities based on SKYLITE_3 analysis, because of missing MS2 data for some isomer of the sum composition 
	# In this case the original total MS1 integral is summed up again for the sum composition and then an MS1 based splitting is applied to generate the quantities of the 
	# (chromatographically resolved) isomers
	if mousemodel==5:
		remergelist=['CR_42:2_(18:2_24:0)', 'CR_42:2_(18:1_24:1)', 'CR_43:1_(19:1_24:0)', 'CR_43:1_(18:1_25:0)']
	elif mousemodel==6:
		remergelist=['CR_42:2_(18:2_24:0)', 'CR_42:2_(18:1_24:1)', 'CR_43:1_(19:1_24:0)', 'CR_43:1_(18:1_25:0)']
	elif mousemodel==7:
		remergelist=['CR_42:2_(18:2_24:0)', 'CR_42:2_(18:1_24:1)', 'CR_43:1_(19:1_24:0)', 'CR_43:1_(18:1_25:0)']
	elif mousemodel==8:
		remergelist=[]
	elif mousemodel==9:
		remergelist=[]
	# begin read Skyline reports file neg
	ntrdf=pd.read_csv('Skyline_Report_isomers_for_SKYLITE_6_pos.csv', low_memory=False)		# SKYLINE REPORT ISOMER RESOLVED IS NOW INPUT FILE
	toprown=[ntrdf.columns.values.tolist()]
	toprown=toprown[0]
	ntrdf=ntrdf.transpose()
	negreport=ntrdf.values.tolist()
	#print('Number of rows in Skyline_Report_JPM_ILS_quantification_neg.csv: %d' % nki)
	# end read  Skyline reports files

	outfilename='SKYLITE_5_quantities_for_SKYLITE_6_pos.xlsx'	# OUTPUT FILE FROM PREVIOUS STEP IS NOW INPUT FILE
	wb=openpyxl.load_workbook(outfilename)
	ws=wb['Quantification']
	ws2=wb.create_sheet('Quantification_isomers')

############################################################################################################################
############################################################################################################################
investigatelipid='SM_33:2_(15:0_18:2)__'
investigatereplicate='NIST1__'
############################################################################################################################
############################################################################################################################
posdataname=negdataname #['NIST1', 'NIST2', 'NIST3']
goff=0					################################# GENERAL OFF SWITCH FOR ALL EXCEPT BLANK ### TURN TO 0 FOR USE OF ALL
chk=1
testrun=0 # default 0, as replicate information needs to be loaded	# IF MANUAL USAGE FOR LOI: SET testrun=1
checkup=0
after=str(beforeall)
tnow=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+' / '+after[11]+after[12]+after[13]+after[14]+after[15]+after[16]+after[17]+after[18]
print('Calculation start at %s' % tnow)


################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
iabh=0.999885
iabc=0.9893
iabn=0.996
iabo=0.99636
iabp=1
###############################################################################################################################


# go through negreport to assign, which entry / index in list belongs to an isomer that is isomer 1, 2, 3...
ixlist=[]			# ixlist indicates if a lipid is only 1 isomer or isomer 1 or 2 or 3, ixlist[ix] is parallel to negreport[0][ix]
ix=1
ixlist.append(ix)
ixi=1
while ixi<len(negreport[0]):
	if negreport[1][ixi]==negreport[1][ixi-1]:
		if negreport[6][ixi]==negreport[6][ixi-1]:
			ix=ix
		elif str(negreport[6][ixi])=='precursor':
			ix=ix+1
		else:
			ix=ix
	else:
		ix=1
	ixlist.append(ix)
	#if ix>0:
	#	if '20_CDDA_2_' in str(negreport[20][ixi]):
	#		print(str(negreport[1][ixi])+'_'+str(ix))
	ixi=ixi+1 
#quit()

if qcheck==2:
	print(ixlist)
	quit()

# search for lipid that requires splitting of existing quantity over chromatographically resolved isomers

#go through previous report and write toprow for new sheet, 
# then go through lipids and apply splitting where required (splitting of quantities done based on MS2 integrals, exceptions noted in ms1splitlist)
# lipids in ms1splitlist will be quantified based on the isomer-associated ms1 integral

ws2.cell(row=1, column=1).value='Lipid'		# ws2 is output sheet (Quantification isomers)
ws2.cell(row=1, column=2).value='RT [min]'
ccol=3
while (ccol-3)<len(negdataname):
	if tissuetype==1:
		repname='c_'+str(negdataname[ccol-3])+' [nmol/mg]' #'replicate '+str(ccol-1)
	else:
		repname='c_'+str(negdataname[ccol-3])+' [nmol/mL]' #'replicate '+str(ccol-1)
	ws2.cell(row=1, column=ccol).value=repname
	ccol=ccol+1

t=2	# index to read previous excel sheet
w=2	# index to write in new excel sheet
#lsl=0
go=1
while go==1:		# go through input data (Quantification)
	#lsl=lsl+1
	clipid=ws.cell(row=t, column=1).value		# ws is input sheet (Quantification)
	clipid=str(clipid)
	slipid=clipid
	isocount=0
	isorlist=[]
	isortlist=[]
	clipidname=str(clipid)
	r=0
	while r<len(negreport[0]):
		if 'PC-O' in str(clipid):
			clipid=clipid[4:]
			clipid='OC'+clipid
		elif 'PC-O/P' in str(clipid):
			clipid=clipid[6:]
			clipid='QC'+clipid
		elif 'PE-O' in str(clipid):
			clipid=clipid[4:]
			clipid='OE'+clipid
		elif 'PE-O/P' in str(clipid):
			clipid=clipid[6:]
			clipid='QE'+clipid
		elif 'LPC' in str(clipid):
			clipid=clipid[3:]
			clipid='LC'+clipid
		elif 'LPE' in str(clipid):
			clipid=clipid[3:]
			clipid='LE'+clipid	
		elif 'LPI' in str(clipid):
			clipid=clipid[3:]
			clipid='LI'+clipid
		elif 'HexCer' in str(clipid):
			cutlipid=clipidname[6:]
			clipid='HC'+cutlipid
		elif 'Cer' in str(clipid):
			cutlipid=clipidname[3:]
			clipid='CR'+cutlipid
			#print(clipid)
		#slipid=clipid
		#elif 'SM' in clipidname:
		#	clipid=clipidname[0]+clipidname[1]+clipidname[2]+clipidname[3]+clipidname[4]+clipidname[5]+clipidname[6]
		#elif 'DG' in clipidname:
		#	clipid=clipidname[0]+clipidname[1]+clipidname[2]+clipidname[3]+clipidname[4]+clipidname[5]+clipidname[6]
		#elif 'TG' in clipidname:
		#	clipid=clipidname[0]+clipidname[1]+clipidname[2]+clipidname[3]+clipidname[4]+clipidname[5]+clipidname[6]
		idcl=0
		if str(clipid) in negreport[1][r]:	#
			idcl=1			
		elif str(clipid)==negreport[1][r]:	#
			idcl=1	
		if idcl==1:
			if str(negreport[6][r])=='precursor':
				if negdataname[0] in str(negreport[20][r]):
					isorlist.append(r)			# list with indeces r in negreport of precursor of rep 1 for each isomer
					isortlist.append(str(negreport[12][r]))
					crt=str(negreport[12][r])
					if isocount>0:
						ok=1
						#print(clipid)
						#print(slipid)
						#print('#############')
					isocount=isocount+1
					# get MS2 quantity to add to cms2sumlist for splitting if required
					
		r=r+1
	if qcheck==1:
		print(clipid)
		print(isocount)
	if 'TG' in clipid:	#ignore multiple TG entries in SKyline report and keep sum composition quantity determined by SKYLITE_5
		isocount=1
	if isocount>1:				# multiple isomers found (e.g., multiple PC(16:0_20:3) species found)
		repi=0
		while repi<len(negdataname):
			
			# check, if species is in ms1splitlist, then split by MS1, otherwise split by MS2
			gosplit=0
			if clipid in ms1splitlist:
				gosplit=1
			if 'SM' in clipid:
				gosplit=1
			if gosplit==1:
				if clipid==investigatelipid:
					if negdataname[repi]==investigatereplicate:
						print('---')
						print(clipid)
						print(negdataname[repi])
						print('# Isomers found: %d' % len(isorlist))

				#if clipid in remergelist:
				#	# re-merge before splitting by MS1
	 			#	ok=1 


				# split by MS1																			### CHECK steps
				# get MS1 integrals of this lipid
				#print(str(negdataname[repi]))
				cms1sumlist=[]
				z=0
				while z<len(isorlist):
					r=isorlist[z]
					crt=isortlist[z]
					cms1sum=0
					cist=negreport[18][r]	# current isomer start time
					ciet=negreport[19][r]	# current isomer end time
					#set s to make sure to catch correct precursor
					s=r
					sgo=1
					while sgo==1:
						if 'precursor' in str(negreport[6][s]):
							sgo=1
							s=s-1
						else:
							sgo=0
					if clipid==investigatelipid:
						if negdataname[repi]==investigatereplicate:
							print('###')
							print('r = %d' % r)
							print('RT Start = %s' % cist)
							print('RT End = %s' % ciet)
					gofa=1
					while gofa==1:
						if 'precursor' in str(negreport[6][s]):
							if str(negdataname[repi]) in str(negreport[20][s]):
								if clipid==investigatelipid:
									if negdataname[repi]==investigatereplicate:
										print('...')
										print('s = %d' % s)
										#print(negreport[20][s])
										print('Precursor integral = %.2f ' % negreport[13][s])
								cms1sum=cms1sum+float(negreport[13][s])
						if abs(cist-negreport[18][s])<0.04:
							if abs(ciet-negreport[19][s])<0.04:
								gofa=gofa
							else:
								gofa=1
						else:
							gofa=1
						if str(negreport[6][s])=='precursor':
							if str(negdataname[repi]) in str(negreport[20][s]):
								gofa=0
						s=s+1
					cms1sumlist.append(cms1sum)
					z=z+1
				# split and write results
				if clipid==investigatelipid:
					if negdataname[repi]==investigatereplicate:
						print(cms1sumlist)


				# begin search for nan entry in cms1sumlist and replace with 0 if found (approximate missing data (isomer quantity) with 0 as not detected, 
				# likely due to low abundance)
				gdl=0
				while gdl<len(cms1sumlist):
					if str(cms1sumlist[gdl])=='nan':
						#dli=cms1sumlist.index('nan')
						cms1sumlist[gdl]=0
					gdl=gdl+1
 

				if clipid==investigatelipid:
					if negdataname[repi]==investigatereplicate:
						print('AFTER ADJUSTMENT:')
						print(cms1sumlist)
						#quit()


				#print(cms2sumlist)
				z=0
				while z<len(isorlist):
					cquant=ws.cell(row=t, column=(repi+3)).value
					wlipid=slipid+'_isomer_'+str(z+1)
					# begin exception for 4 mousemodels
					if wlipid=='SM_40:2_(22:0_18:2)_isomer_1':
						wlipid='SM_40:2_isomer_1'
					elif wlipid=='SM_40:2_(22:0_18:2)_isomer_2':
						wlipid='SM_40:2_isomer_2'
					elif wlipid=='SM_40:2_(22:0_18:2)_isomer_3':
						wlipid='SM_40:2_isomer_3'
					elif wlipid=='DG_36:2_(18:1_18:1)_isomer_1':
						wlipid='DG_36:2_(18:1_18:1)'
					elif wlipid=='DG_36:2_(18:1_18:1)_isomer_2':
						wlipid='DG_36:2_(18:0_18:2)'
					elif wlipid=='DG_38:6_(18:2_20:4)_isomer_1':
						wlipid='DG_38:6_(18:2_20:4)'
					elif wlipid=='DG_38:6_(18:2_20:4)_isomer_2':
						wlipid='DG_38:6_(16:0_22:6)'
					# end exception for 4 mousemodels
					ws2.cell(row=w+z, column=1).value=wlipid
					crt=isortlist[z]
					ws2.cell(row=w+z, column=2).value=crt			# RT
					if sum(cms1sumlist)==0:
						wquant=cquant
					elif cquant is None:
						wquant=cquant
					else:
						#print(cquant)
						#print(cms1sumlist[z])
						#print((sum(cms1sumlist)))
						wquant=cquant*(cms1sumlist[z]/sum(cms1sumlist))
					if wquant==0:
						wquant=''
					else:
						ws2.cell(row=w+z, column=(repi+3)).value=wquant
					z=z+1
				repi=repi+1

			else:
				# get MS2 integrals of all isomer features of this lipid
				#print(str(negdataname[repi]))
				cms2sumlist=[]
				z=0
				while z<len(isorlist):
					r=isorlist[z]
					crt=isortlist[z]
					cms2sum=0

					parked=1			# old module requiring RT match between replicates
					if parked==0:
						cist=negreport[18][r]	# current isomer start time
						ciet=negreport[19][r]	# current isomer end time
						gofa=1
						s=r+1+repi
						while gofa==1:
							if 'FA' in str(negreport[6][s]):
								if str(negdataname[repi]) in str(negreport[20][s]):
									cms2sum=cms2sum+float(negreport[13][s])
							elif 'LCB' in str(negreport[6][s]):
								if str(negdataname[repi]) in str(negreport[20][s]):
									cms2sum=cms2sum+float(negreport[13][s])
							if abs(cist-negreport[18][s])<0.04:
								if abs(ciet-negreport[19][s])<0.04:
									gofa=gofa
								else:
									gofa=1
							else:
								gofa=1
							if str(negreport[6][s])=='precursor':
								if str(negdataname[repi]) in str(negreport[20][s]):
									gofa=0
							s=s+1

					# new module based on ixlist for assignment of isomer number
					grb=0
					while grb<len(negreport[1]):
						if str(negdataname[repi]) in str(negreport[20][grb]):	# replicate name
							if negreport[1][grb]==clipid:					# lipid name
								if (z+1)==ixlist[grb]:						# isomer number
									if 'FA' in str(negreport[6][grb]):		# FA diagnostic fragment
										cms2sum=cms2sum+float(negreport[13][grb])
									elif 'LCB' in str(negreport[6][grb]):		# FA diagnostic fragment
										cms2sum=cms2sum+float(negreport[13][grb])
						grb=grb+1
					
					cms2sumlist.append(cms2sum)
					z=z+1
				# split and write results
				#print(cms2sumlist)
				z=0
				while z<len(isorlist):
					cquant=ws.cell(row=t, column=(repi+3)).value
					wlipid=slipid+'_isomer_'+str(z+1)
					ws2.cell(row=w+z, column=1).value=wlipid
					crt=isortlist[z]
					ws2.cell(row=w+z, column=2).value=crt			# RT
					if sum(cms2sumlist)==0:
						wquant=cquant
					elif cquant is None:
						wquant=cquant
					else:
						#print(cquant)
						#print(cms2sumlist[z])
						#print((sum(cms2sumlist)))
						wquant=cquant*(cms2sumlist[z]/sum(cms2sumlist))
					if wquant==0:
						wquant=''
					else:
						ws2.cell(row=w+z, column=(repi+3)).value=wquant
					z=z+1
				repi=repi+1
		w=w+len(isorlist)
	
	else:  
		# exception for 4 mousemodels
		if 'SM' in slipid:
			wslipid=slipid[0]+slipid[1]+slipid[2]+slipid[3]+slipid[4]+slipid[5]+slipid[6]
		else:
			wslipid=slipid
		ws2.cell(row=w, column=1).value=wslipid	# copy lipid name and RT without any splitting or modification
		ws2.cell(row=w, column=2).value=crt
		repi=0
		while repi<len(negdataname):
			# check for remergelist to correct splitting ???
			if clipid in remergelist:
				#
				#print('....')
				#print(clipid)
				#print(slipid)
				#print(negdataname[repi])
				# retrieve original sum of quantities for the current sum composition lipid (e.g., abs quantity of sum of all PC 38:7 isomers)
				# go through excel file
				originalsum=0
				ccl=repi+3
				crw=2
				csc=clipid[0]+clipid[1]+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]
				cscs=slipid[0]+slipid[1]+slipid[2]+slipid[3]+slipid[4]+slipid[5]+slipid[6]+slipid[7]
				gorw=1
				while gorw==1:
					selipid=ws.cell(row=crw, column=1).value
					if selipid is None:
						gorw=0
					elif str(selipid)=='':
						gorw=0
					else:
						if cscs in selipid:
							oquant=ws.cell(row=crw, column=ccl).value
							originalsum=originalsum+oquant
							gorw=0
					crw=crw+1
				#print(originalsum)
				# get MS1 precursor integral of all lipid isomers within sum composition as well as current lipid
				cms1list=[]
				cms1i=0
				sr=0
				while sr<len(negreport[0]):
					if csc in str(negreport[1][sr]):
						if 'precursor'==str(negreport[6][sr]):
							if str(negdataname[repi]) in str(negreport[20][sr]):
								cms1list.append(negreport[13][sr])
								if clipid==negreport[1][sr]:
									cms1i=negreport[13][sr]
					sr=sr+1
				# calc fraction of MS1 precursor intensities and apply to original sum composition quantity, then write in file
				nq=cms1i/sum(cms1list)*originalsum
				#print(cms1i)
				#print(cms1list)
				#print(nq)
				if nq==0:
					nq=''
				else:
					ws2.cell(row=w, column=repi+3).value=nq
				
			else:
				cquant=ws.cell(row=t, column=(repi+3)).value
				if cquant==0:
					cquant=''
				else:
					ws2.cell(row=w, column=(repi+3)).value=cquant
			repi=repi+1
		w=w+1
	t=t+1
	clipid=ws.cell(row=t, column=1).value
	if clipid is None:
		go=0
	elif str(clipid)=='NaN':
		go=0
	elif str(clipid)=='':
		go=0
	else:
		go=1




after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'

runidentifier=wsin3.cell(row=2, column=3).value
runidentifier=str(runidentifier)
if today in runidentifier:
	outfilename=runidentifier+'_SKYLITE_Q6_pos.xlsx'
else:
	outfilename=today+runidentifier+'_SKYLITE_Q6_pos.xlsx'

#outfilename=outfilename[:-10]
#outfilename=today+'SKYLITE_6_isomer_quantities_pos.xlsx'
wb.save(outfilename)
#################################################################################################################################################################
#################################################################################################################################################################
if excelinput==1:
	print('Calculation completed. The output file with isomer resolved quantities is saved as: %s' % outfilename)
	quit()


	
print('Calculating data for volcano plots...')

#################################################################################################################################################################
#################################################################################################################################################################
wb=openpyxl.load_workbook(outfilename)
ws2=wb['Quantification_isomers']

# indeces for data for volcano plots [CCl4, MC4R-KO, STAM, CDDA] - indeces of column in excel file Quantification_isomers
vstartctrli=[3, 14, 27, 39]
#vendctrli=[7, 19, 32, 43]
#vstartmashi=[8, 20, 33, 44]
#vendmashi=[12, 25, 38, 48]
vsnum=[5, 6, 6, 5]	# number of replicates per mouse model

tvolc=4
volci=0
while volci<tvolc:
	print('...')
	lplist=[]	# lipid list
	pvlist=[]	# raw p value
	ablist=[]	# mean abundance of lipid in set of replicates with higher abundance
	lgfclist=[]	# log2 fold change list
	# Begin calculate means, standard deviations, fold changes and p-values
	snum=vsnum[volci] #repnumvolcano #6	# sample number for statistics calculation (e.g., comparison of 6 WT with 6 KO samples, these have to be the first 6 (12) samples in list negdataname)
	csnum=snum
	go=1
	r=2
	while go==1:
		csnum=snum
		clipid=ws2.cell(row=r, column=1).value
		asum=0
		bsum=0
		alist=[]
		blist=[]
		sni=vstartctrli[volci] #0
		while sni<(snum+vstartctrli[volci]):
			caq=ws2.cell(row=r, column=sni).value
			if caq is None:
				caq=0
				csnum=csnum-1
			else:
				cbq=ws2.cell(row=r, column=sni+snum).value
				if cbq is None:
					cbq=0
					csnum=csnum-1
				else:
					caq=float(caq)
					asum=asum+caq
					alist.append(caq)
					cbq=float(cbq)
					bsum=bsum+cbq
					blist.append(cbq)
			sni=sni+1
		if csnum==0:
			amean=0
			bmean=0
		else:
			amean=asum/csnum
			bmean=bsum/csnum
		#print('...')
		#print(clipid)
		#print(alist)
		#print(blist)

		# begin shorten clipid name to remove sum composition from name of DG, Cer and HexCer, shorten isomer to i 
		clipid=str(clipid)
		if 'isomer' in clipid:
			ii=clipid[len(clipid)-1]
			clipid=clipid[:-7]+ii
		if 'DG' in clipid:
			ni=0
			while ni<len(clipid):
				if clipid[ni]==':':
					clipid=clipid[:-(len(clipid)-ni+3)]+clipid[-((len(clipid)-ni)-3):]
					ni=len(clipid)
				ni=ni+1
		elif 'Cer' in clipid:
			ni=0
			while ni<len(clipid):
				if clipid[ni]==':':
					clipid=clipid[:-(len(clipid)-ni+3)]+clipid[-((len(clipid)-ni)-4):]
					ni=len(clipid)
				ni=ni+1
			clipid=clipid[:-1]
			ni=0
			while ni<len(clipid):
				if clipid[ni]=='_':
					clipid=clipid[:-(len(clipid)-ni)]+';2/'+clipid[-((len(clipid)-ni-1)):]
					ni=len(clipid)
				ni=ni+1
			if clipid=='Cer22:0;2/18:0':
				clipid='Cer_40:0'
		elif 'HexCer' in clipid:
			ni=0
			while ni<len(clipid):
				if clipid[ni]==':':
					clipid=clipid[:-(len(clipid)-ni+6)]+clipid[-((len(clipid)-ni)-5):]
					ni=len(clipid)
				ni=ni+1

		# end shorten clipid name to remove sum composition from name of PC, PE and PI, shorten isomer to i 
		vci=23*volci
		ws2.cell(row=r, column=3+len(negdataname)+2+vci).value=clipid
		ws2.cell(row=r, column=3+len(negdataname)+3+vci).value=amean    #write 
		ws2.cell(row=r, column=3+len(negdataname)+5+vci).value=bmean    #write 
		if len(alist)<2:
			astd=''
		else:
			astd=statistics.stdev(alist)
		if len(blist)<2:
			bstd=''
		else:
			bstd=statistics.stdev(blist)
		ws2.cell(row=r, column=3+len(negdataname)+4+vci).value=astd    #write 
		ws2.cell(row=r, column=3+len(negdataname)+6+vci).value=bstd    #write 
		if amean==0:
			fcm=0
		else:
			fcm=bmean/amean
			ws2.cell(row=r, column=3+len(negdataname)+7+vci).value=fcm    #write 
		#print(fcm)
		if fcm==0:
			lgfcm=0
		else:
			lgfcm=math.log2(fcm)
			ws2.cell(row=r, column=3+len(negdataname)+9+vci).value=lgfcm    #write 
		lgfclist.append(lgfcm)
		if lgfcm==0:
			ok=1
			# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
			lplist.append(clipid)
			ablist.append(0)
			pvlist.append(1)
			# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
		if lgfcm==0:
			ok=1
		else:
			# begin calc P value
			t,p=stats.ttest_ind(alist, blist, equal_var=False)
			ts=str(t)
			ps=str(p)
			if ts=='nan':
				t=str(1.0)
			if ps=='nan':
				p=str(1.0)
			t=float(t)
			p=float(p)
			#print(p)
			#print(t)
			# write P value and ttest statistic in excel output file
			ws2.cell(row=r, column=3+len(negdataname)+8+vci).value=p    #write 
			ws2.cell(row=r, column=3+len(negdataname)+12+vci).value=t    #write 
			clgp=math.log10(p)	# p value for volcano plot as -log10()
			clgp=(-1)*clgp
			sign=0
			if p<0.05:		# determine if fold change and p value significant
				if fcm>2:
					sign=1
				elif fcm<0.5:
					sign=1
			if sign==1:
				ws2.cell(row=r, column=3+len(negdataname)+11+vci).value=clgp    #write significant p values / with significant fold changes
			else:
				if clgp==0:
					ok=1
				else:
					ws2.cell(row=r, column=3+len(negdataname)+10+vci).value=clgp    #write insignificant
			# end calc P value
			# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
			lplist.append(clipid)
			if amean>bmean:
				lmean=amean
			else:
				lmean=bmean
			ablist.append(lmean)
			pvlist.append(p)
			# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down

		r=r+1
		clipid=ws2.cell(row=r, column=1).value
		if clipid is None:
			go=0
		elif str(clipid)=='NaN':
			go=0
		elif str(clipid)=='':
			go=0
		else:
			go=1


	ws2.cell(row=1, column=3+len(negdataname)+2+vci).value='Lipid'
	ws2.cell(row=1, column=3+len(negdataname)+3+vci).value='Mean [col C-H]'
	ws2.cell(row=1, column=3+len(negdataname)+4+vci).value='Std. dev. [col C-H]'
	ws2.cell(row=1, column=3+len(negdataname)+5+vci).value='Mean [col I-N]'
	ws2.cell(row=1, column=3+len(negdataname)+6+vci).value='Std. dev. [col I-N]'
	ws2.cell(row=1, column=3+len(negdataname)+7+vci).value='Fold change of mean'
	ws2.cell(row=1, column=3+len(negdataname)+8+vci).value='P value'
	ws2.cell(row=1, column=3+len(negdataname)+9+vci).value='log2 fold change'
	ws2.cell(row=1, column=3+len(negdataname)+10+vci).value='="-log10 p-value (not significant)"'
	ws2.cell(row=1, column=3+len(negdataname)+11+vci).value='="-log10 p-value (significant)"'
	ws2.cell(row=1, column=3+len(negdataname)+12+vci).value='t-test statistic'




	# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
	ws2.cell(row=1, column=3+len(negdataname)+13+vci).value='P value (Bonferroni)'
	ws2.cell(row=1, column=3+len(negdataname)+14+vci).value='P value (BH)'
	ws2.cell(row=1, column=3+len(negdataname)+15+vci).value='P value (ASD)'
	ws2.cell(row=1, column=3+len(negdataname)+16+vci).value='="-log10 p-value (Bonferroni)"'
	ws2.cell(row=1, column=3+len(negdataname)+17+vci).value='="-log10 p-value (ASD)"'
	ws2.cell(row=1, column=3+len(negdataname)+18+vci).value='="-log10 p-value (BH; insignificant)"'
	ws2.cell(row=1, column=3+len(negdataname)+19+vci).value='="-log10 p-value (BH; also significant after ASD and BF)"'
	ws2.cell(row=1, column=3+len(negdataname)+20+vci).value='="-log10 p-value (BH; also significant after ASD, but not BF)"'
	ws2.cell(row=1, column=3+len(negdataname)+21+vci).value='="-log10 p-value (BH; significant, but not after ASD or BF)"'
	nmc=len(lplist)

	#print('Lists for BH:')
	# begin create sorted lists by p value for BH
	#print(lplist)
	#print(pvlist)
	psortlplist=[]
	psortpvlist=[]
	dlplist=[]
	dpvlist=[]
	ni=0
	while ni<len(lplist):
		dlplist.append(lplist[ni])
		dpvlist.append(pvlist[ni])
		ni=ni+1
	while len(dlplist)>0:
		cpv=min(dpvlist)
		cdi=dpvlist.index(cpv)
		psortpvlist.append(dpvlist[cdi])
		psortlplist.append(dlplist[cdi])
		del dpvlist[cdi]
		del dlplist[cdi]
	#print(lplist)
	#print(pvlist)
	#print(psortlplist)
	#print(psortpvlist)
	# end create sorted lists by p value for BH
	#print('Lists for ASD:')
	# begin create sorted lists by p value for ASD
	#print(lplist)
	#print(pvlist)
	#print(ablist)
	absortlplist=[]
	absortpvlist=[]
	absortablist=[]
	dlplist=[]
	dpvlist=[]
	dablist=[]
	ni=0
	while ni<len(lplist):
		dlplist.append(lplist[ni])
		dpvlist.append(pvlist[ni])
		dablist.append(ablist[ni])
		ni=ni+1
	while len(dlplist)>0:
		cab=min(dablist)
		cdi=dablist.index(cab)
		absortpvlist.append(dpvlist[cdi])
		absortablist.append(dablist[cdi])
		absortlplist.append(dlplist[cdi])
		del dpvlist[cdi]
		del dlplist[cdi]
		del dablist[cdi]
	#print(lplist)
	#print(pvlist)
	#print(absortlplist)
	#print(absortpvlist)
	#print(absortablist)
	# end create sorted lists by p value for ASD
	apvi=0
	while apvi<len(lplist):
		if pvlist[apvi]==1:
			ok=1
		elif pvlist[apvi] is None:
			ok=1
		else:
			# begin Bonferroni
			bfpv=pvlist[apvi]*nmc
			if bfpv>1:
				bfpv=1
			logbfpv=math.log10(bfpv)	# p value for volcano plot as -log10()
			logbfpv=(-1)*logbfpv
			ws2.cell(row=2+apvi, column=3+len(negdataname)+13+vci).value=bfpv
			ws2.cell(row=2+apvi, column=3+len(negdataname)+16+vci).value=logbfpv
			# end Bonferroni

			# begin Abundance-Step-Down
			kasd=(absortlplist.index(lplist[apvi]))
			asdpv=pvlist[apvi]*(nmc-kasd)
			if asdpv>1:
				asdpv=1
			logasdpv=math.log10(asdpv)	# p value for volcano plot as -log10()
			logasdpv=(-1)*logasdpv
			ws2.cell(row=2+apvi, column=3+len(negdataname)+14+vci).value=asdpv
			ws2.cell(row=2+apvi, column=3+len(negdataname)+17+vci).value=logasdpv
			# end Abundance-Step-Down

			# begin Benjamini-Hochberg
			kbh=(psortlplist.index(lplist[apvi]))+1
			bhpv=pvlist[apvi]*(nmc/kbh)
			if bhpv>1:
				bhpv=1
			logbhpv=math.log10(bhpv)	# p value for volcano plot as -log10()
			logbhpv=(-1)*logbhpv
			ws2.cell(row=2+apvi, column=3+len(negdataname)+15+vci).value=bhpv
			# determine colour code (column), depending on: 
			# 1) Insignificant after BH (18)
			# 2) Significant after BH and significant after BF (19)
			# 3) Significant after BH and significant after ASD, but not after BF (20)
			# 4) Significant after BH, but not after ASD or BF (21)
			if abs(lgfclist[apvi])<1:
				ws2.cell(row=2+apvi, column=3+len(negdataname)+18+vci).value=logbhpv	# 1) Insignificant due to small fold change
			else:
				if bhpv>0.05:
					ws2.cell(row=2+apvi, column=3+len(negdataname)+18+vci).value=logbhpv	# 1) Insignificant after BH
				else:
					if bfpv<0.05:
						ws2.cell(row=2+apvi, column=3+len(negdataname)+19+vci).value=logbhpv	# 2) Significant after BH and significant after BF and ASD
					else:
						if asdpv<0.05:
							ws2.cell(row=2+apvi, column=3+len(negdataname)+20+vci).value=logbhpv	# 3) Significant after BH and significant after ASD, but not after BF
						else:
							ws2.cell(row=2+apvi, column=3+len(negdataname)+21+vci).value=logbhpv	# 4) Significant after BH, but not after ASD or BF 
			# end Benjamini-Hochberg

		apvi=apvi+1
	# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down

	volci=volci+1

# End calculate means, standard deviations, fold changes and p-values

outfilename=today+'SKYLITE_6_isomer_quantities_pos.xlsx'
wb.save(outfilename)

print('Calculation completed. The output file with isomer resolved quantities is saved as: %s' % outfilename)
#print('Manually correct PC_34:3_(16:0_18:3)_isomer1 ALA in MC4R-KO_HCD !!')



