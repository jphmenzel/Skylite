# -*- coding: UTF-8 -*-

# Jan Philipp Menzel
# Performs: Reading reports from Skyline files, positive mode, quantification based on internal standards at sum composition level
# Note: This version performs isotope correction type one based on sum formula
# First version created on 6/12/2023
## Notes: Output file will be named: today+'SKYLITE_3_quantities_for_SKYLITE_4_neg.xlsx'
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import subprocess
import statistics
import csv
import sys
maxInt=sys.maxsize
beforeall=datetime.datetime.now()

print('--------------------------------------------------------------------------------------------------------')
print(' S K Y L I T E  5 ')
print('This program calculates quantities of lipids from curated Skyline report files of the SKYLITE workflow (positive polarity).')
print('The required input file for this program is Skyline_Report_curated_for_SKYLITE_5_pos.csv')
print('Please refer to the associated publication / preprint and the github page for information.')
print('For each run of this program, adjust the replicate naming pattern and the replicate weights for correct quantification.')
print('--------------------------------------------------------------------------------------------------------')
print('########### C H E C K L I S T ####################')
print('   1) Set tissue type, line 39')
print('   2) Check liver homogenate volume, line 40')
print('   3) Check internal standard amount, line 43')
print('   4) Set mousemodel, line 51')
print('   5) Edit elif clause for mousemodel, containing sample names, liver tissue weight and liver protein amount, after line 172')
print('   6) If required, enter rule for internal standard integration, line 494')
calculate=eval(input('All edited according to checklist? Y=1, N=0 ___'))
if calculate==0:
	quit()
################################################# MODIFY naming pattern of replicates
tissuetype=1		# liver tissue = 1;  blood plasma = 0	# MODIFY for sample type
vollivhom=20		# 20 is 20uL of liver homogenate that was used for lipid extraction												

# define used Ultimate SPLASH volumes ! ####################														# MODIFY
uspvol=1*0.001*0.001 # L (1 uL)			volume of original Ultimate SPLASH solution used	#default is eq. of 1 uL Ultimate SPLASH used for each sample
samplevol=10*0.001*0.001 # L (10 uL)	sample volume of NIST human plasma or bioliquid				# MODIFY, use 10 uL in case of liver tissue analysis, see below
if tissuetype==1:
	samplevol=10*0.001*0.001		# use 10 uL in case of liver tissue analysis
# define used Ultimate SPLASHvolumes ! ####################

# Select mouse model ######################################
#negdataname=['1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', '7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', '13_CTRL_1_', '14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', '20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', '24b_CDDA_6_', '24c_CDDA_6_', 'NIST_QC1_', 'NIST_QC2_']		# naming pattern of replicates
mousemodel=8 # 6 is NIST validation 	# 5 is all 4 together				# MODIFY TO SELECT WHICH VOLCANO PLOT IS TO BE MADE


if mousemodel==1:
	# CCl4 HFD mouse model Inselspital ###########################################
	negdataname=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
			  'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_', 
			  'NIST_QC1_i1', 'NIST_QC1_i1_2', 'PBQC_CML', 'PBQC_MML']
	replicateweights=[12.6, 12.8, 14.6, 14.2, 12.1, 
				   13.8, 12.6, 14.1, 14.8, 11.8, 
				   10, 10, 13, 13]
	
elif mousemodel==2:
	# MC4R-KO_HCD mouse model Deborah Stroka ######################################
	negdataname=['FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
				 'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
				 'FB4_', 'FB20_', 'FB5_', 'FB6_', 'FB7_', 'FB8_', 'FB17_', 'FB18_', 'FB19_', 'NIST_QC1_', 'NIST_QC2_']		# naming pattern of replicates
	replicateweights=[15.08, 14.36, 14.46, 9.59, 16.58, 10.80, 
					  14.50, 17.58, 15.41, 15.11, 11.17, 17.51, 
					  10.83, 18.42, 16.98, 10.95, 12.82, 15.56, 18.40, 16.80, 13.49, 10, 10]

elif mousemodel==3:
	# Choline deficient CDDA mouse model ######################################
	negdataname=['14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
				'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
				'24b_CDDA_6_', '24c_CDDA_6_', 
				'1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
				'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
				'NIST_QC1_', 'NIST_QC2_', '13_CTRL_1_']		# naming pattern of replicates
	# use replicateweights for liver tissue, mass in mg of liver tissue used for homogenization, assuming standard protocol (check below for vol of liver homogenate used)
	replicateweights=[12.27, 18.55, 20.78, 20.72, 20.70, 
					18.13, 20.21, 14.66, 13.00, 17.73, 
					17.73, 17.73, 
					14.40, 10.21, 10.98, 11.65, 16.08, 13.86, 
					15.01, 10.90, 15.42, 15.81, 18.88, 18.49, 
					10, 10, 20.63]
elif mousemodel==4:
	# Streptozotocin diabetic STAM mouse model ######################################
	negdataname=['1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
				'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
				'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
				'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
				'24b_CDDA_6_', '24c_CDDA_6_', 
				'NIST_QC1_', 'NIST_QC2_', '13_CTRL_1_']		# naming pattern of replicates
	# use replicateweights for liver tissue, mass in mg of liver tissue used for homogenization, assuming standard protocol (check below for vol of liver homogenate used)
	replicateweights=[14.40, 10.21, 10.98, 11.65, 16.08, 13.86, 
					15.01, 10.90, 15.42, 15.81, 18.88, 18.49,
					12.27, 18.55, 20.78, 20.72, 20.70, 
					18.13, 20.21, 14.66, 13.00, 17.73, 
					17.73, 17.73,
					10, 10, 20.63]
elif mousemodel==5:
	# CCl4 HFD mouse model Inselspital ###########################################
	nd1=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
		'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_', 
		'USP_pos_NIST_QC1_i1']
	rw1=[12.6, 12.8, 14.6, 14.2, 12.1, 
		13.8, 12.6, 14.1, 14.8, 11.8, 
		10]
	rpw1=[1.40, 2.17, 2.43, 1.94, 1.94, 
		1.92, 1.64, 1.71, 1.83, 1.20, 
		1, 1]
	# MC4R-KO_HCD mouse model Deborah Stroka ######################################
	nd2=['FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
		'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
		'DML_pos_NIST_QC1_']		# naming pattern of replicates
	rw2=[15.08, 14.36, 14.46, 9.59, 16.58, 10.80, 
		14.50, 17.58, 15.41, 15.11, 11.17, 17.51, 
		10]
	rpw2=[2.36, 2.32, 2.30, 1.20, 3.25, 1.67, 
		2.20, 2.66, 2.10, 2.11, 1.26, 2.76, 1]
	# Choline deficient CDDA mouse model ######################################
	# Streptozotocin diabetic STAM mouse model ######################################
	nd3=['1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
		'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
		'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
		'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
		'24b_CDDA_6_', '24c_CDDA_6_', 'KML_pos_NIST_QC2_']		# naming pattern of replicates
	# use replicateweights for liver tissue, mass in mg of liver tissue used for homogenization, assuming standard protocol (check below for vol of liver homogenate used)
	rw3=[14.40, 10.21, 10.98, 11.65, 16.08, 13.86, 
		15.01, 10.90, 15.42, 15.81, 18.88, 18.49,
		12.27, 18.55, 20.78, 20.72, 20.70, 
		18.13, 20.21, 14.66, 13.00, 17.73, 
		17.73, 17.73, 10]
	rpw3=[2.57, 1.78, 1.76, 1.85, 3.29, 2.09,  
		1.87, 1.30, 2.00, 1.64, 2.09, 2.62,  
		1.92, 3.84, 3.73, 3.85, 3.89,  
		1.87, 2.51, 1.35, 1.13, 1.84, 
		1.84, 1.84, 1]
	negdataname=[]
	replicateweights=[]
	repproteinweight=[]
	ndi=0
	while ndi<len(nd1):
		negdataname.append(nd1[ndi])
		replicateweights.append(rw1[ndi])
		repproteinweight.append(rpw1[ndi])
		ndi=ndi+1
	ndi=0
	while ndi<len(nd2):
		negdataname.append(nd2[ndi])
		replicateweights.append(rw2[ndi])
		repproteinweight.append(rpw2[ndi])
		ndi=ndi+1
	ndi=0
	while ndi<len(nd3):
		negdataname.append(nd3[ndi])
		replicateweights.append(rw3[ndi])
		repproteinweight.append(rpw3[ndi])
		ndi=ndi+1
elif mousemodel==6:
	negdataname=['NIST1_', 'NIST2_', 'NIST3_']
elif mousemodel==7:			# liver interday validation
	negdataname=['KML_', 'KML24d_', 'KML24e_']
	replicateweights=[17.73, 17.73, 17.73]
	repproteinweight=[1.84, 1.84, 1.84]
elif mousemodel==8:
	# MC4R-KO_HCD mouse model Deborah Stroka ######################################
	negdataname=['FB1_', 'FB2_', 'FB3_', 'NIST_QC1_']		# naming pattern of replicates
	replicateweights=[14.50, 17.58, 15.41, 10]
	repproteinweight=[2.20, 2.66, 2.10, 1]
# use replicatevolume for blood plasma, volume in microL of plasma
#replicatevolume=[]

################################################# MODIFY 

after=str(beforeall)
tnow=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+' / '+after[11]+after[12]+after[13]+after[14]+after[15]+after[16]+after[17]+after[18]
print('Calculation start at %s' % tnow)

###############################################################################################################################
posdataname=negdataname #['NIST1', 'NIST2', 'NIST3']
goff=0					################################# GENERAL OFF SWITCH FOR ALL EXCEPT BLANK ### TURN TO 0 FOR USE OF ALL
chk=1
testrun=0 # default 0, as replicate information needs to be loaded	# IF MANUAL USAGE FOR LOI: SET testrun=1
checkup=0
# begin read Skyline reports file pos
#trdf=pd.read_csv('Skyline_Report_JPM_ILS_quantification_pos.csv')
#trdf=pd.read_csv('Skyl_Rep_JPM_NIST_USP_pos_curated.csv')
#toprowp=[trdf.columns.values.tolist()]
#toprowp=toprowp[0]
#trdf=trdf.transpose()
#posreport=trdf.values.tolist()
#print('Number of rows in Skyline_Report_JPM_ILS_quantification_pos.csv: %d' % ki)
# begin read Skyline reports file neg
ntrdf=pd.read_csv('Skyline_Report_curated_for_SKYLITE_5_pos.csv', low_memory=False)
toprown=[ntrdf.columns.values.tolist()]
toprown=toprown[0]
ntrdf=ntrdf.transpose()
negreport=ntrdf.values.tolist()
#print('Number of rows in Skyline_Report_JPM_ILS_quantification_neg.csv: %d' % nki)
# end read  Skyline reports files

################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
iabh=0.999885
iabc=0.9893
iabn=0.996
iabo=0.99636
iabp=1
################


# begin definition of Ultimate SPLASH according to transition list and used volumes and amounts
#uspindexref=[0, 1, 2, 3, 4,			PC
#			 5, 6, 7, 8, 9,				PE
#			 10, 11, 12, 13, 14,		PG
#			 15, 16, 17, 18, 19,		PI
#			 20, 21, 22,				LPC
#			 23, 24, 25,				LPE
#			 26, 27, 28,				LPG
#			 29, 30, 31,				LPI
#			 32, 33, 34, 35, 36,		CM
#			 37, 38, 39, 40, 41,		SM
#			 42, 43, 44, 45, 46,		CE
#			 47, 48, 49, 50, 51,		DG
#			 52, 53, 54, 				TG
#			 55, 56, 57,				TG
#			 58, 59, 60]				TG
uspname=['PC 14:1-17:0 (d5)', 'PC 16:1-17:0 (d5)', 'PC 17:0-18:1 (d5)', 'PC 17:0-20:3 (d5)', 'PC 17:0-22:4 (d5)',
		 'PE 14:1-17:0 (d5)', 'PE 16:1-17:0 (d5)', 'PE 17:0-18:1 (d5)', 'PE 17:0-20:3 (d5)', 'PE 17:0-22:4 (d5)',
		 'PG 14:1-17:0 (d5)', 'PG 16:1-17:0 (d5)', 'PG 17:0-18:1 (d5)', 'PG 17:0-20:3 (d5)', 'PG 17:0-22:4 (d5)',
		 'PI 14:1-17:0 (d5)', 'PI 16:1-17:0 (d5)', 'PI 17:0-18:1 (d5)', 'PI 17:0-20:3 (d5)', 'PI 17:0-22:4 (d5)',
		 'LPC 15:0 (d5)', 'LPC 17:0 (d5)', 'LPC 19:0 (d5)', 
		 'LPE 15:0 (d5)', 'LPE 17:0 (d5)', 'LPE 19:0 (d5)', 
		 'LPG 15:0 (d5)', 'LPG 17:0 (d5)', 'LPG 19:0 (d5)', 
		 'LPI 15:0 (d5)', 'LPI 17:0 (d5)', 'LPI 19:0 (d5)', 
		 'Cer 18:1;2/16:1 (d7)', 'Cer 18:1;2/18:1 (d7)', 'Cer 18:1;2/20:1 (d7)', 'Cer 18:1;2/22:1 (d7)', 'Cer 18:1;2/24:1 (d7)',
		 'SM 18:1;2/16:1 (d9)', 'SM 18:1;2/18:1 (d9)', 'SM 18:1;2/20:1 (d9)', 'SM 18:1;2/22:1 (d9)', 'SM 18:1;2/24:1 (d9)',
		 'SE 27:1/14:1 (d7)', 'SE 27:1/16:1 (d7)', 'SE 27:1/18:1 (d7)', 'SE 27:1/20:3 (d7)', 'SE 27:1/22:4 (d7)', 
		 'DAG 14:1-17:0 (d5)', 'DAG 16:1-17:0 (d5)', 'DAG 17:0-18:1 (d5)', 'DAG 17:0-20:3 (d5)', 'DAG 17:0-22:4 (d5)',
		 'TAG 13:0-14:0-14:0 (d5)', 'TAG 14:0-14:0-15:1 (d5)', 'TAG 14:0-14:0-17:1 (d5)', 
		 'TAG 15:1-16:0-16:0 (d5)', 'TAG 16:0-16:0-17:1 (d5)', 'TAG 16:0-16:0-19:2 (d5)', 
		 'TAG 17:1-18:1-18:1 (d5)', 'TAG 18:1-18:1-19:2 (d5)', 'TAG 18:1-18:1-21:2 (d5)']
uspc=[50, 100, 150, 100, 50,
	  25, 50, 75, 50, 25,
	  25, 50, 75, 50, 25,
	  25, 50, 75, 50, 25,
	  25, 50, 25,
	  25, 50, 25,
	  25, 50, 25,
	  25, 50, 25,
	  75, 50, 25, 50, 75,
	  75, 50, 25, 50, 75,
	  25, 50, 75, 50, 25,
	  25, 50, 75, 50, 25,
	  25, 50, 75,
	  100, 125, 100,
	  75, 50, 25]		# concentration of original Ultimate SPLASH standards in ug/mL (mg/L)
uspmm=[723.04, 751.09, 779.15, 803.17, 829.21,
	   680.96, 709.01, 737.07, 761.09, 787.13,
	   733.95, 762.01, 790.06, 814.08, 840.12,
	   817.06, 845.12, 873.17, 897.19, 923.23,
	   486.63, 514.7, 542.75,
	   444.55, 472.61, 500.67,
	   497.55, 525.61, 553.66,
	   580.66, 608.72, 636.77,
	   542.93, 570.98, 599.03, 627.09, 655.14,
	   710.08, 738.14, 766.19, 794.24, 822.28,
	   602.04, 630.09, 658.14, 682.18, 708.2,
	   557.91, 585.97, 614.02, 638.04, 664.08,
	   714.18, 740.22, 768.27,
	   796.33, 824.38, 850.42,
	   876.46, 902.47, 930.53]	# molar mass of Ultimate SPLASH standards in g/mol
uspfracformula=['C39H71NO8P', 'C41H75NO8P', 'C43H79NO8P', 'C45H79NO8P', 'C47H81NO8P', 
				'C36H65NO8P', 'C38H69NO8P', 'C40H73NO8P', 'C42H73NO8P', 'C44H75NO8P', 
				'C37H65O10P', 'C39H69O10P', 'C41H73O10P', 'C43H73O10P', 'C45H75O10P', 
				'C40H69O13P', 'C42H73O13P', 'C44H77O13P', 'C46H77O13P', 'C48H79O13P', 
				'C23H43NO7P', 'C25H47NO7P', 'C27H51NO7P', 
				'C20H37NO7P', 'C22H41NO7P', 'C24H45NO7P', 
				'C21H37O9P', 'C23H41O9P', 'C25H45O9P', 
				'C24H41O12P', 'C26H45O12P', 'C28H49O12P', 
				'C34H58NO3', 'C36H62NO3', 'C38H66NO3', 'C40H70NO3', 'C42H74NO3', 
				'C39H68N2O6P', 'C41H72N2O6P', 'C43H76N2O6P', 'C45H80N2O6P', 'C47H84N2O6P', 
				'C41H63O2', 'C43H67O2', 'C45H71O2', 'C47H71O2', 'C49H73O2', 
				'C34H59O5', 'C36H63O5', 'C38H67O5', 'C40H67O5', 'C42H69O5', 
				'C44H79O6', 'C46H81O6', 'C48H85O6', 
				'C50H89O6', 'C52H93O6', 'C54H95O6', 
				'C56H97O6', 'C58H99O6', 'C60H103O6']		# fractional sum formula of Ultimate SPLASH standards (without d atoms)


uspn=[]		# molar amount of original Ultimate SPLASH standards used 
uspi=0
while uspi<len(uspname):
	cuspn=(uspvol*uspc[uspi]*0.001)/uspmm[uspi]	#current n [mol] of USP standard used in sample
	uspn.append(cuspn)
	uspi=uspi+1
uspcs=[]	# USP concentration in unknown sample
uspi=0
while uspi<len(uspname):
	cuspcs=uspn[uspi]/samplevol	#current c [mol/L] of USP standard equivalent in sample
	uspcs.append(cuspcs)
	uspi=uspi+1
# begin definition of Ultimate SPLASH according to transition list and used volumes and amounts

def assignment_model(clipidname):
	ui=0
	cergo=0
	if 'P'==clipidname[0]:
		plgo=1
	elif 'O'==clipidname[0]:
		plgo=1
	elif 'Q'==clipidname[0]:
		plgo=1
	else:
		plgo=0
	if plgo==1:
		if int(clipidname[6])<2:
			if int(clipidname[4])+(10*int(clipidname[3]))<32:
				ui=0
			elif int(clipidname[4])+(10*int(clipidname[3]))<35:
				ui=1
			else:
				ui=2
		elif int(clipidname[6])<4:
			ui=3
		else:
			ui=4
		if 'E'==clipidname[1]:
			ui=ui+5
		elif 'G'==clipidname[1]:
			ui=ui+10
		elif 'I'==clipidname[1]:
			ui=ui+15
	elif 'L'==clipidname[0]:
		if int(clipidname[4])+(10*int(clipidname[3]))<16:
			ui=20
		elif int(clipidname[4])+(10*int(clipidname[3]))<19:
			ui=21
		else:
			ui=22
		if 'E'==clipidname[1]:
			ui=ui+3
		elif 'G'==clipidname[1]:
			ui=ui+6
		elif 'I'==clipidname[1]:
			ui=ui+9
		ok=1
	elif 'CR' in clipidname:
		cergo=1
	elif 'HC' in clipidname:
		cergo=1
	elif 'DC' in clipidname:
		cergo=1
	elif 'SM' in clipidname:
		cergo=1
	elif 'CE' in clipidname:
		if int(clipidname[6])<2:
			if int(clipidname[4])+(10*int(clipidname[3]))<15:
				ui=42
			elif int(clipidname[4])+(10*int(clipidname[3]))<18:
				ui=43
			else:
				ui=44
		elif int(clipidname[6])<4:
			ui=45
		else:
			ui=46
	elif 'DG' in clipidname:
		if int(clipidname[6])<2:
			if int(clipidname[4])+(10*int(clipidname[3]))<33:
				ui=47
			elif int(clipidname[4])+(10*int(clipidname[3]))<35:
				ui=48
			else:
				ui=49
		elif int(clipidname[6])<4:
			ui=50
		else:
			ui=51
	elif 'TG' in clipidname:
		if int(clipidname[6])<2:
			if int(clipidname[4])+(10*int(clipidname[3]))<42:
				ui=52			# TG_41:0_(13:0_14:0_14:0)_d5 
			elif int(clipidname[4])+(10*int(clipidname[3]))<44:
				ui=53			# TG_43:1_(14:0_14:0_15:1)_d5 
			elif int(clipidname[4])+(10*int(clipidname[3]))<47:
				ui=54			# TG_45:1_(14:0_14:0_17:1)_d5 
			elif int(clipidname[4])+(10*int(clipidname[3]))<49:
				ui=55			# TG_47:1_(15:1_16:0_16:0)_d5 
			else:
				ui=56			# TG_49:1_(16:0_16:0_17:1)_d5
		elif int(clipidname[6])<3:
			ui=57				# TG_51:2_(16:0_16:0_19:2)_d5 
		elif int(clipidname[6])<4:
			ui=58				# TG_53:3_(17:1_18:1_18:1)_d5 
		else:
			if int(clipidname[4])+(10*int(clipidname[3]))<57:
				ui=59  			# TG_55:4_(18:1_18:1_19:2)_d5 
			else:
				ui=60			# TG_57:4_(18:1_18:1_21:2)_d5 
	if cergo==1:	# Ceramides
		if int(clipidname[4])+(10*int(clipidname[3]))<35:
			ui=32				# Cer 18:1;2/16:1 
		elif int(clipidname[4])+(10*int(clipidname[3]))<37:
			ui=33				# Cer 18:1;2/18:1 
		elif int(clipidname[4])+(10*int(clipidname[3]))<40:
			ui=34				# Cer 18:1;2/20:1 
		elif int(clipidname[4])+(10*int(clipidname[3]))<42:
			ui=35				# Cer 18:1;2/22:1 
		else:
			ui=36				# Cer 18:1;2/24:1 
	if 'SM' in clipidname:
		ui=ui+5
	# begin enter any exclusion of specific IS due to unreliable IS quantification ( if ui==xx, then ui=zy)
	if mousemodel<6:	# 4 mousemodels
		if ui==52:
			ui=53 		# TG_41:0_(13:0_14:0_14:0)_d5 is unreliable, use reliable TG_43:1_(14:0_14:0_15:1)_d5 instead
		elif ui==56:
			ui=55 		# TG_49:1_(16:0_16:0_17:1)_d5 is unreliable, use reliable TG_47:1_(15:1_16:0_16:0)_d5 instead
		elif ui==57:
			ui=55 		# TG_51:2_(16:0_16:0_19:2)_d5 is unreliable, use reliable TG_47:1_(15:1_16:0_16:0)_d5 instead
		elif ui==58:
			ui=55 		# TG_53:3_(17:1_18:1_18:1)_d5 is unreliable, use reliable TG_47:1_(15:1_16:0_16:0)_d5 instead
		elif ui==59:
			ui=55 		# TG_55:4_(18:1_18:1_19:2)_d5 is unreliable, use reliable TG_47:1_(15:1_16:0_16:0)_d5 instead
		elif ui==60:
			ui=55 		# TG_57:4_(18:1_18:1_21:2)_d5 is unreliable, use reliable TG_47:1_(15:1_16:0_16:0)_d5 instead
		elif ui==33:
			ui=34 		# Cer_18:1;2/18:1_d7 is unreliable, use reliable Cer_18:1;2/20:1_d7 instead
		elif ui==39:
			ui=40 		# SM_18:1;2/20:1_d9 is unreliable, use reliable SM_18:1;2/20:1_d9 instead
		elif ui==47:
			ui=50 		# DG_(14:1_17:0)_d5 is unreliable, use reliable DG_(17:0_20:3)_d5 instead
		elif ui==48:
			ui=50 		# DG_(16:1_17:0)_d5 is unreliable, use reliable DG_(17:0_20:3)_d5 instead
		elif ui==49:
			ui=50 		# DG_(17:0_18:1)_d5 is unreliable, use reliable DG_(17:0_20:3)_d5 instead
	elif mousemodel==6:		# Plasma NIST validation
		if ui==33:
			ui=34 		# Cer_18:1;2/18:1_d7 is unreliable, use reliable Cer_18:1;2/20:1_d7 instead
		elif ui==48:
			ui=49 		# DG_(16:1_17:0)_d5 is unreliable, use reliable DG_(17:0_20:3)_d5 instead
	elif mousemodel==7:
		if ui==60:
			ui=59
		elif ui==57:
			ui=56
		elif ui==58:
			ui=56
		elif ui==32:
			ui=36
		elif ui==33:
			ui=36
		elif ui==34:
			ui=36
		elif ui==35:
			ui=36
		elif ui==38:
			ui=37
		elif ui==47:
			ui=51
		elif ui==48:
			ui=51
		elif ui==49:
			ui=51
		elif ui==50:
			ui=51
	elif mousemodel==8:
		if ui==52:
			ui=52
		elif ui==54:
			ui=54
	# end enter any exclusion of specific IS due to unreliable IS quantification ( if ui==xx, then ui=zy)
	uspindex=ui
	return uspindex

#print(negreport[1][3])
#print(negreport[2][5])

# lipid species is in column index=1 of report file, e.g., PE_33:1_(15:0_18:1d7) 
# lipid transition (precursor etc) is in column index=6
# mzError (mass accuracy, error in ppm) is in column index=11
# RT (min) is in column index=12
# Integral (a.u.) is in column index=13
# sample name (dataset file name, e.g., 2023_11_28_JPM_ILS_neg_HTG_F1a.raw) is in column index=20

def isotopecorrect(pcformula, pcintis):
	#correct integral of lipid from monoisotopic ion to full isotopic envelope, read cformula and cintis
	#print(pcformula)
	#print(pcintis)

	e=str(pcformula) #
	# begin read precursor sum formula and edit product sum formula
	#print(e)
	#print(e[0])
	#print(e[1])
	#print(e[2])
	#print(e[3])
	#print(len(e))
	clist=[]
	hlist=[]
	dlist=[]
	nlist=[]
	olist=[]
	plist=[]
	ilist=[]
	i=0
	ca=0
	ha=0
	da=0
	na=0
	oa=0
	pa=0
	ia=0
	while i<len(e):
		if e[i]=='H':
			if e[i+1]=="'":
				ha=0
			else:
				ca=0
		#if e[i]=="'":
		#	ha=0		
		if e[i]=='N':
			ha=0
			da=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
		if e[i]=='P':
			ha=0
			da=0
			na=0
			oa=0
		if e[i]=='I':
			ha=0
			da=0
			na=0
			oa=0
			pa=0
		if ca==1:
			clist.append(e[i])
		if ha==1:
			hlist.append(e[i])
		if da==1:
			dlist.append(e[i])
		if na==1:
			nlist.append(e[i])
		if oa==1:
			olist.append(e[i])
		if pa==1:
			plist.append(e[i])
		if ia==1:
			ilist.append(e[i])
		if e[i]=='C':
			ca=1
		if e[i]=='H':
			if e[i+1]=="'":
				ca=0
				ha=0
				da=1
				i=i+1
			else:
				ca=0
				ha=1
		#if e[i]=="'":
		#	ca=0
		#	ha=0
		#	da=1		
		if e[i]=='N':
			ha=0
			da=0
			na=1
			if e[i+1]=='O':
				nlist.append('1')
				na=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
			oa=1
			if (i+1)<len(e):
				if e[i+1]=='P':
					olist.append('1')
					oa=0
			else:
				olist.append('1')
				oa=0					
		if e[i]=='P':
			da=0
			na=0
			oa=0
			pa=1
			if (i+1)<len(e):
				if e[i+1]=='I':
					plist.append('1')
					pa=0
			else:
				plist.append('1')
				pa=0
		if e[i]=='I':
			da=0
			na=0
			oa=0
			pa=0
			ia=1
			if i==(len(e)-1):
				ilist.append('1')
				ia=0
		i=i+1
	#print(clist)
	#print(hlist)
	#print(dlist)
	#print(nlist)
	#print(olist)
	#print(plist)
	if len(clist)==0:
		cn=0
	if len(hlist)==0:
		hn=0
	if len(dlist)==0:
		dn=0	
	if len(nlist)==0:
		nn=0
	if len(olist)==0:
		on=0
	if len(plist)==0:
		pn=0
	if len(ilist)==0:
		iodon=0
	if len(clist)==1:
		cn=int(clist[0])
	if len(clist)==2:
		cn=10*int(clist[0])+int(clist[1])
	if len(clist)==3:
		cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
	if len(hlist)==1:
		hn=int(hlist[0])
	if len(hlist)==2:
		hn=10*int(hlist[0])+int(hlist[1])
	if len(hlist)==3:
		hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
	if len(dlist)==1:
		dn=int(dlist[0])
	if len(dlist)==2:
		dn=10*int(dlist[0])+int(dlist[1])
	if len(dlist)==3:
		dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
	if len(nlist)==1:
		nn=int(nlist[0])
	if len(nlist)==2:
		nn=10*int(nlist[0])+int(nlist[1])
	if len(nlist)==3:
		nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
	if len(olist)==1:
		on=int(olist[0])
	if len(olist)==2:
		on=10*int(olist[0])+int(olist[1])
	if len(olist)==3:
		on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
	if len(plist)==1:
		pn=int(plist[0])
	if len(plist)==2:
		pn=10*int(plist[0])+int(plist[1])
	if len(plist)==3:
		pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
	if len(ilist)==1:
		iodon=int(ilist[0])
	if len(ilist)==2:
		iodon=10*int(ilist[0])+int(ilist[1])
	if len(ilist)==3:
		iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		
	# end read precursor sum formula
	if exa==1:
		print(cn)
		print(hn)
		print(nn)
		print(on)
		print(pn)
	iabh=0.999885
	iabc=0.9893
	iabn=0.996
	iabo=0.99636
	iabp=1
	monoiso=iabh**hn*iabc**cn*iabo**on*iabn**nn*iabp**pn		# fraction of the isotopic pattern being the monoisotopic ion
	outcintis=pcintis/monoiso
	#outcintis=pcintis #*2
	return outcintis

exa=0
if exa==1:
	cformula='C12H20N3O4P'
	cintis=100
	outcintis=isotopecorrect(cformula, cintis)
	print(cformula)
	print(outcintis)
	quit()


###############################################################################################################
###############################################################################################################
###############################################################################################################
# begin calculate quantification values, write results in report sheet

# Notes: Only use neg mode for PI !!!
# Use only NH4 ionization for CE !!!

after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'



wb = Workbook(write_only=True)
outfilename=today+'SKYLITE_5_quantities_for_SKYLITE_6_pos.xlsx'
wb.save(outfilename)
wb=openpyxl.load_workbook(outfilename)
ws=wb.create_sheet('Quantification')
del wb['Sheet']
#ws=wb.active

#ws.cell(row=1, column=1).value='Lipid'
ws.cell(row=1, column=1).value='Lipid'
ws.cell(row=1, column=2).value='RT [min]'
ccol=3
while (ccol-3)<len(negdataname):
	if tissuetype==1:
		repname='c_'+str(negdataname[ccol-3])+' [nmol/mg]' #' [nmol/mL]' for plasma, and ' [nmol/mg]' for liver tissue			# MODIFY if change between liver tissue and plasma
	else:
		repname='c_'+str(negdataname[ccol-3])+' [nmol/mL]' #' [nmol/mL]' for plasma, and ' [nmol/mg]' for liver tissue			# MODIFY if change between liver tissue and plasma
	ws.cell(row=1, column=ccol).value=repname
	ccol=ccol+1
ws.cell(row=1, column=3+len(negdataname)+6).value='Associated IS'	# name of internal standard used for this lipid
# go through negreport, then posreport to gather integrals of lipids and IS, use IS according to quantification model to quantify

#print(negreport[0][0])
#print(negreport[0][1])
#print(negreport[2][0])
#print(len(negreport[0]))

# begin write list of lipids
lplist=[]
lpilist=[]
r=0
count=2
while r<len(negreport[0]):
	if 'd5' in str(negreport[1][r]):
		ok=1
	elif 'd7' in str(negreport[1][r]):
		ok=1
	elif 'd9' in str(negreport[1][r]):
		ok=1
	elif str(negreport[6][r])=='precursor':
		if str(negdataname[0]) in str(negreport[20][r]):
			clipidname=str(negreport[1][r])
			if 'OC' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='PC-O'+cutlipid
			elif 'QC' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='PC-O/P'+cutlipid
			elif 'OE' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='PE-O'+cutlipid
			elif 'QE' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='PE-O/P'+cutlipid
			elif 'LC' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='LPC'+cutlipid
			elif 'LE' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='LPE'+cutlipid
			elif 'LG' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='LPG'+cutlipid
			elif 'LI' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='LPI'+cutlipid
			elif 'HC' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='HexCer'+cutlipid
			elif 'CR' in clipidname:
				cutlipid=clipidname[2:]
				clipidname='Cer'+cutlipid
			elif 'TG' in clipidname:
				clipidname=clipidname[0]+clipidname[1]+clipidname[2]+clipidname[3]+clipidname[4]+clipidname[5]+clipidname[6]
			lplist.append(clipidname)
			lpilist.append(count)
			count=count+1
	r=r+1
# end write list of lipids

#print(lplist)
#quit()

crow=1
wrlist=[]

rnnf=1
r=0
while r<len(negreport[0]):
	if 'd5' in str(negreport[1][r]):
		ok=1
	elif 'd7' in str(negreport[1][r]):
		ok=1
	elif 'd9' in str(negreport[1][r]):
		ok=1
	elif str(negreport[6][r])=='precursor':
		rnnf=1
		#print(negreport[20][r])
		sid=0
		while sid<len(negdataname):
			if str(negdataname[sid]) in str(negreport[20][r]):		# check if names in negdataname ok
				cid=negdataname[sid]		# current replicate or sample name according to negdataname as defined above (part of sample or replicate name)
				cidi=sid					# current replicate or sample index according to negdataname
				rnnf=0
			sid=sid+1
	
		#if rnnf==1:
		#	#print(negreport[20][r])
		if rnnf==0:
			#print(negdataname[cidi])
			clipidname=str(negreport[1][r])
			# check, if integral shall be used as is or splitting according to MS2 required for isomer integration
			splitint=0
			k=0
			while k<len(negreport[0]):
				if str(negreport[1][k])==str(negreport[1][r]):
					ok=1
				elif negreport[4][k]==negreport[4][r]:	# found an isomer of the same sum composition, but different FAs
					if negreport[0][k]==negreport[0][r]:
						if str(negreport[6][k])=='precursor':
							if negreport[20][k]==negreport[20][r]:		# same replicate
								# determine, if all integration limits separate from each other, allowing individual MS1 based integration (splitint=0) or if integral needs splitting based on MS2
								if negreport[18][k]>negreport[19][r]:	# col 18 is start time of integral, col 19 is end time
									splitint=0
								elif negreport[18][r]>negreport[19][k]:
									splitint=0
								else:
									splitint=1	# found at least one isomer with different FA composition, which requires splitting integral according to MS2
									# get MS2 integrals of current r and all k isomers
									# count number of isomers and find indices for all relevant MS2 integrals
									ms2sumlist=[]	
									isoindexlist=[]	
									isoindexlist.append(r)
									isoindexlist.append(k)
									# get MS2 for r, then k, then look for others
									s=r
									cms2int=0
									while str(negreport[1][s])==str(negreport[1][r]):
										if 'FA' in negreport[6][s]:
											if negreport[20][s]==negreport[20][r]:
												cms2int=cms2int+float(negreport[13][s])
										s=s+1
									ms2sumlist.append(cms2int)	# [r_ms2int]
									s=k
									cms2int=0
									while str(negreport[1][s])==str(negreport[1][k]):
										if 'FA' in negreport[6][s]:
											if negreport[20][s]==negreport[20][k]:
												cms2int=cms2int+float(negreport[13][s])
										s=s+1
									ms2sumlist.append(cms2int)	# [r_ms2int, k_ms2int]
									# search for other isomers in sum composition
									t=0
									while t<len(negreport[0]):
										if str(negreport[1][t])==str(negreport[1][r]):
											ok=1
										elif str(negreport[1][t])==str(negreport[1][k]):
											ok=1
										else:
											if negreport[4][t]==negreport[4][r]:
												if negreport[0][t]==negreport[0][r]:
													if str(negreport[6][t])=='precursor':
														if negreport[20][t]==negreport[20][r]:
															s=t
															cms2int=0
															while str(negreport[1][s])==str(negreport[1][t]):
																if 'FA' in negreport[6][s]:
																	if negreport[20][s]==negreport[20][t]:
																		cms2int=cms2int+float(negreport[13][s])
																s=s+1
															ms2sumlist.append(cms2int)	# [r_ms2int, k_ms2int, t1_ms2int, ...]
															isoindexlist.append(t)
										t=t+1
									# proceed with splitting integrals by MS2
									isoi=0
									while isoi<len(ms2sumlist):
										# do isotope correction
										clipidname=str(negreport[1][isoindexlist[isoi]])
										crt=str(negreport[12][r])
										cint=negreport[13][r]
										cformula=negreport[2][r]
										uspindex=assignment_model(clipidname) #index in usp lists to define the selected internal standard
										cformulausp=uspfracformula[uspindex]		# get formula of applied internal standard USP, so that isotope correction can be applied

										cint=isotopecorrect(cformula, cint)
										# identify appropriate standard to use
										# assignment model
										# get integral of respective internal standard
										h=0
										while h<len(negreport[0]):
											if str(uspname[uspindex])==(negreport[1][h]):
												if str(negreport[6][h])=='precursor':
													if str(negreport[20][r])==str(negreport[20][h]):
														cintagainst=negreport[13][h]
														h=len(negreport[0])
											h=h+1
										cintagainst=isotopecorrect(cformulausp, cintagainst)		# Isotope correction internal standard USP
										apuspname=uspname[uspindex]	# applied USP, name
										apuspcs=uspcs[uspindex]		# applied USP, respective concentration
										cconc=(cint/cintagainst)*apuspcs	# concentration of lipid
										if 'OC' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='PC-O'+cutlipid
										elif 'QC' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='PC-O/P'+cutlipid
										elif 'OE' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='PE-O'+cutlipid
										elif 'QE' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='PE-O/P'+cutlipid
										elif 'LC' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='LPC'+cutlipid
										elif 'LE' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='LPE'+cutlipid
										elif 'LG' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='LPG'+cutlipid
										elif 'LI' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='LPI'+cutlipid
										elif 'HC' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='HexCer'+cutlipid
										elif 'CR' in clipidname:
											cutlipid=clipidname[2:]
											clipidname='Cer'+cutlipid
										elif 'TG' in clipidname:
											clipidname=clipidname[0]+clipidname[1]+clipidname[2]+clipidname[3]+clipidname[4]+clipidname[5]+clipidname[6]

										if clipidname in lplist:
											# search for correct row to add to
											rrow=lplist.index(clipidname)
											crow=rrow+2
											#print(clipidname)
											#print(crow)
										else:
											print('############# CHECK #################')
											crow=500+crow+len(isoindexlist)						###			CHECK INDEXING !!!!!!!!!!!!!!

										ws.cell(row=crow, column=1).value=clipidname
										ws.cell(row=crow, column=2).value=crt
										if float(sum(ms2sumlist))==0:
											wconc=cconc*1000*1000
											#print(clipidname)
											#print(negdataname[cidi])
										else:
											wconc=cconc*(ms2sumlist[isoi]/sum(ms2sumlist))*1000*1000  # concentration of lipid in sample solution in nmol/mL

										nlpd=wconc*(0.01)	# nlpd is n_lipid in sample used for lipid extraction in nmol (assuming above noted 10 uL plasma)
										# above noted calculation compensated for earlier entry of 10 uL plasma, to allow in the following to calculated lipid amounts per mg liver tissue
										# use replicateweights list to calculate nmol/mg of liver tissue
										#vollivhom=20	# 20 is 20uL of liver homogenate that was used for lipid extraction							# MODIFY IF REQUIRED ################# 
										totalvollivhom=1000	# 1000 is 1000 uL of liver homogenate generated initially from liver tissue
										# determine vollivhom dependent on current sample type
										if tissuetype==1:
											if 'CML' in negdataname[cidi]:
												vollivhom=10
											elif 'MML' in negdataname[cidi]:
												vollivhom=10
											else:
												vollivhom=20
											nlpdliver=(nlpd/vollivhom)*totalvollivhom	# nlpdliver is amount n in nmol of lipid in the total of the liver homogenate (in the 1 mL)
											crepweight=replicateweights[cidi]		# crepweight is the liver tissue weight for this current replicate
											crepliverw=nlpdliver/crepweight		# cperliverw is the amount n in nmol of lipid per mg of liver tissue

											if 'NIST' in negdataname[cidi]:
												crepquantnorm=wconc #crepquantnorm=crepliverw
											else:
												crepprotein=repproteinweight[cidi]	# crepprotein is the amount of protein in mg in the piece of liver tissue that was used for homogenization
												crepquantnorm=crepliverw/(crepprotein/crepweight)	# crepquantnorm is the amount if lipid in nmol per mg of liver protein 

										if tissuetype==1:	# 1 is liver tissue
											if crepquantnorm==0:
												ok=1
											else:
												ws.cell(row=crow, column=3+cidi).value=crepquantnorm
										else:
											ws.cell(row=crow, column=3+cidi).value=wconc
										if cidi==0:
											ws.cell(row=crow, column=3+len(negdataname)+6+cidi).value=apuspname
										if clipidname in wrlist:
											ok=1
										else:
											wrlist.append(clipidname)
										isoi=isoi+1

								#negreport[13][j]
				k=k+1
			if splitint==0:
				# do isotope correction
				crt=str(negreport[12][r])
				cint=negreport[13][r]
				cformula=negreport[2][r]

				# identify appropriate standard to use
				# assignment model
				uspindex=assignment_model(clipidname) #index in usp lists to define the selected internal standard

				cformulausp=uspfracformula[uspindex]		# get formula of applied internal standard USP, so that isotope correction can be applied

				cint=isotopecorrect(cformula, cint)							# Isotope correction analyte
				
				# get integral of respective internal standard
				h=0
				while h<len(negreport[0]):
					if str(uspname[uspindex])==(negreport[1][h]):
						if str(negreport[6][h])=='precursor':
							if str(negreport[20][r])==str(negreport[20][h]):
								cintagainst=negreport[13][h]
								h=len(negreport[0])
					h=h+1

				cintagainst=isotopecorrect(cformulausp, cintagainst)		# Isotope correction internal standard USP

				apuspname=uspname[uspindex]	# applied USP, name
				apuspcs=uspcs[uspindex]		# applied USP, respective concentration
				if cintagainst==0:
					cconc=0
				else:
					cconc=(cint/cintagainst)*apuspcs	# concentration of lipid
					if 'TG_52:2_edit' in clipidname:
						print(negdataname[cidi])
						print(cint)
						print(cintagainst)
						print(apuspcs)
						print(apuspname)
						print(uspindex)
						print(cconc)
						print('......')
				if 'OC' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='PC-O'+cutlipid
				elif 'QC' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='PC-O/P'+cutlipid
				elif 'OE' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='PE-O'+cutlipid
				elif 'QE' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='PE-O/P'+cutlipid
				elif 'LC' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='LPC'+cutlipid
				elif 'LE' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='LPE'+cutlipid
				elif 'LG' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='LPG'+cutlipid
				elif 'LI' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='LPI'+cutlipid
				elif 'HC' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='HexCer'+cutlipid
				elif 'CR' in clipidname:
					cutlipid=clipidname[2:]
					clipidname='Cer'+cutlipid
				elif 'TG' in clipidname:
					clipidname=clipidname[0]+clipidname[1]+clipidname[2]+clipidname[3]+clipidname[4]+clipidname[5]+clipidname[6]
				if clipidname in lplist:
					# search for correct row to add to
					rrow=lplist.index(clipidname)
					crow=rrow+2
					#print(clipidname)
					#print(crow)
				else:
					print('############# CHECK replicate naming pattern #################')
					crow=crow+1+500

				ws.cell(row=crow, column=1).value=clipidname
				ws.cell(row=crow, column=2).value=crt
				wconc=1000*1000*cconc	# wconc is concentration of lipid in sample solution in nmol/mL (e.g., per 10 uL human plasma)

				nlpd=wconc*(0.01)	# nlpd is n_lipid in sample used for lipid extraction in nmol (assuming above noted 10 uL plasma)
				# above noted calculation compensated for earlier entry of 10 uL plasma, to allow in the following to calculated lipid amounts per mg liver tissue
				# use replicateweights list to calculate nmol/mg of liver tissue
				#vollivhom=20	# 20 is 20uL of liver homogenate that was used for lipid extraction							# MODIFY IF REQUIRED (ABOVE) ################# 
				# determine vollivhom dependent on current sample type
				if 'CML' in negdataname[cidi]:
					vollivhom=10
				elif 'MML' in negdataname[cidi]:
					vollivhom=10
				elif 'NIST' in negdataname[cidi]:
					vollivhom=10
				else:
					vollivhom=20
				if tissuetype==1:
					totalvollivhom=1000	# 1000 is 1000 uL of liver homogenate generated initially from liver tissue
					nlpdliver=(nlpd/vollivhom)*totalvollivhom	# nlpdliver is amount n in nmol of lipid in the total of the liver homogenate (in the 1 mL)
					crepweight=replicateweights[cidi]		# crepweight is the liver tissue weight for this current replicate
					crepliverw=nlpdliver/crepweight		# cperliverw is the amount n in nmol of lipid per mg of liver tissue
					if 'NIST' in negdataname[cidi]:
						crepquantnorm=wconc #crepquantnorm=crepliverw
					else:
						crepprotein=repproteinweight[cidi]	# crepprotein is the amount of protein in mg in the piece of liver tissue that was used for homogenization
						crepquantnorm=crepliverw/(crepprotein/crepweight)	# crepquantnorm is the amount if lipid in nmol per mg of liver protein 
					
				if tissuetype==1:	# 1 is liver tissue
					if crepquantnorm==0:
						ok=1
					else:
						ws.cell(row=crow, column=3+cidi).value=crepquantnorm #crepliverw #crepliverw to be used for liver tissue	normalized to liver weight, crepquantnorm normalized to protein content
				else:
					ws.cell(row=crow, column=3+cidi).value=wconc		#wconc to be used for plasma, crepliverw to be used for liver tissue			
				if cidi==0:
					ws.cell(row=crow, column=3+len(negdataname)+6+cidi).value=apuspname	# name of internal standard used for this lipid
				if clipidname in wrlist:
					ok=1
				else:
					wrlist.append(clipidname)	
	r=r+1

wb.save(outfilename)
print('Calculation completed. The output file is saved as %s' % outfilename)

###################################################################################################################################################
###################################################################################################################################################
###################################################################################################################################################
###################################################################################################################################################







