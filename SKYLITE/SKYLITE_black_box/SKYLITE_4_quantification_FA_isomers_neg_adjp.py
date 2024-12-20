# -*- coding: UTF-8 -*-

# Developer: Dr. Jan Philipp Menzel
# Performs: Reading reports from Skyline adjustment of integrals and duplication for quantification of chromatographically resolved isomers (negative mode)
# Note: reads previous output from SKYLITE 3 to determine how some quantities have to be split based on isomers present 
# First version created on 6/12/2023
## Notes: 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import subprocess
import statistics
import csv
import sys
import scipy
from scipy import stats
maxInt=sys.maxsize
beforeall=datetime.datetime.now()

excelinput=1	# default is the use of SKYLITE_INPUT_samplelist.xlsx as a method of entering metadata into the workflow
if excelinput==1:
	print('--------------------------------------------------------------------------------------------------------')
	print(' S K Y L I T E  4 ')
	print('This program calculates quantities of lipid isomers from curated Skyline report files of the SKYLITE workflow and from the output file of SKYLITE 3.')
	print('The required input Skyline report file has to be named according to SKYLITE_INPUT_samplelist.xlsx')
	print('All input data requested in SKYLITE_INPUT_samplelist.xlsx is required.')
	print('Please refer to the associated publication / preprint and the github page for further information.')
	print('--------------------------------------------------------------------------------------------------------')
	wb=openpyxl.load_workbook('SKYLITE_INPUT_sample_list.xlsx')
	ws1=wb['Sample_list']
	ws2=wb['Ultimate_SPLASH_ONE_exceptions']
	ws3=wb['Input_file_names']
	ws4=wb['Isomer_integration_exceptions']
	internalstandard=ws1.cell(row=2, column=7).value
	tissuetype=ws1.cell(row=2, column=3).value
	vollivhom=ws1.cell(row=2, column=4).value
	uspvolul=ws1.cell(row=2, column=8).value
	uspvol=float(uspvolul)*0.001*0.001	# convert from uL to L
	ismvolul=ws1.cell(row=2, column=9).value
	ismvol=float(ismvolul)*0.001*0.001	# convert from uL to L
	samplevol=10*0.001*0.001	# convert from uL to L
	negdataname=[]
	replicateweights=[]
	repproteinweight=[]
	goread=1
	rrow=2
	while goread==1:
		csample=ws1.cell(row=rrow, column=2).value
		if csample is None:
			goread=0
		elif str(csample)=='nan':
			goread=0
		else:
			negdataname.append(str(csample))
			ctt=ws1.cell(row=2, column=3).value
			if ctt==1:
				vollivhom=ws1.cell(row=rrow, column=4).value
				crepwt=ws1.cell(row=rrow, column=5).value
				if crepwt is None:
					replicateweights.append(10)
				else:
					replicateweights.append(float(crepwt))
				crepprwt=ws1.cell(row=rrow, column=6).value
				if crepprwt is None:
					repproteinweight.append(1)
				else:
					repproteinweight.append(float(crepprwt))
			elif ctt==0:
				csamplevol=ws1.cell(row=rrow, column=4).value
				samplevol=csamplevol*0.001*0.001	# convert from uL to L
				replicateweights.append(10)
				repproteinweight.append(1)
		rrow=rrow+1

	ms1splitlist=[]
	goread=1
	rrow=2
	while goread==1:
		csplit=ws4.cell(row=rrow, column=2).value
		if csplit is None:
			goread=0
		elif str(csplit)=='nan':
			goread=0
		else:
			ms1splitlist.append(str(csplit))
		rrow=rrow+1

	remergelist=[]
	goread=1
	rrow=2
	while goread==1:
		crm=ws4.cell(row=rrow, column=3).value
		if crm is None:
			goread=0
		elif str(crm)=='nan':
			goread=0
		else:
			remergelist.append(str(crm))
		rrow=rrow+1


	cfn=ws3.cell(row=20, column=6).value	#row defines file # reading isomer report from Skyline
	filenameskylrep4=str(cfn)+'.csv'
	ntrdf=pd.read_csv(filenameskylrep4, low_memory=False)
	toprown=[ntrdf.columns.values.tolist()]
	toprown=toprown[0]
	ntrdf=ntrdf.transpose()
	negreport=ntrdf.values.tolist()

	cfn2=ws3.cell(row=23, column=6).value	#row defines file
	filenameq3s4=str(cfn2)+'.xlsx'	# OUTPUT FILE FROM PREVIOUS STEP IS NOW INPUT FILE
	wbq3s4=openpyxl.load_workbook(filenameq3s4)
	ws1q3s4=wbq3s4['Quantification']
	ws2q3s4=wbq3s4.create_sheet('Quantification_isomers')

	snum=0
	repnumvolcano=0
else:
	print('--------------------------------------------------------------------------------------------------------')
	print(' S K Y L I T E  4 ')
	print('This program calculates quantities of lipid isomers from curated Skyline report files of the SKYLITE workflow and from the output file of SKYLITE 3.')
	print('The required input files for this program are:')
	print(' - Skyline_Report_isomers_for_SKYLITE_4_neg.csv')
	print(' - SKYLITE_3_quantities_for_SKYLITE_4_neg.xlsx')
	print('Manually enter or modify species in ms1splitlist, if applicable:')
	print('These are species at low abundance with missing or inconsistent MS2 data, but isomers, which are chromatographically separated.')
	print('Please refer to the associated publication / preprint and the github page for information.')
	print('--------------------------------------------------------------------------------------------------------')
	print('########### C H E C K L I S T ####################')
	print('   1) Set tissue type, line 41')
	print('   4) Set mousemodel, line 42')
	print('   5) Edit elif clause for mousemodel, containing sample names, after line 127')
	print('   6) Edit elif clause for ms1splitlist, containing lipid names for exceptions, after line 120')
	print('   6) Edit elif clause for remergelist, containing lipid names for exceptions, after line 149')
	calculate=eval(input('All edited according to checklist? Y=1, N=0 ___'))
	if calculate==0:
		quit()
	tissuetype=1		# liver tissue = 1;  blood plasma = 0	# MODIFY for sample type
	mousemodel=8		# model 1 = CCl4_HFD # model 2 = MC4R-KO_HFD # model 3 = CDDA_HFD # 4 = STAM  # 7 = liver interday
	################################################# MODIFY 
	#negdataname=['NIST1', 'NIST2', 'NIST3']		# naming pattern of replicates# MODIFY 
	#negdataname=['CML1', 'CML2', 'CML3', 'CML4', 'CML5', 'MML1', 'MML2', 'MML3', 'MML4', 'MML5', 'NIST_QC1_i1', 'NIST_QC1_i1_2', 'PBQC_CML', 'PBQC_MML']		# naming pattern of replicates
	if mousemodel==1: # CCl4_HFD
		repnumvolcano=5	
		negdataname=['CML1_', 'CML2_', 'CML3_', 'CML4_', 'CML5_', 
				'MML1_', 'MML2_', 'MML3_', 'MML4_', 'MML5_', 
				'NIST_QC1_i1', 'NIST_QC1_i1_2'] #, 'PBQC_CML', 'PBQC_MML']
	elif mousemodel==2:	# MC4R-KO_HFD
		repnumvolcano=6
		negdataname=['FB9_', 'FB10_', 'FB11_', 'FB12_', 'FB13_', 'FB21_', 
				'FB1_', 'FB2_', 'FB3_', 'FB14_', 'FB15_', 'FB16_', 
				'FB4_', 'FB20_', 'FB5_', 'FB6_', 'FB7_', 'FB8_', 'FB17_', 'FB18_', 'FB19_', 'NIST_QC1_', 'NIST_QC2_']		# naming pattern of replicates
	elif mousemodel==3: # CDDA HFD
		# Choline deficient:
		repnumvolcano=5	
		negdataname=['14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
				'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
				'24b_CDDA_6_', '24c_CDDA_6_', 
				'1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
				'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
				'NIST_QC1_', 'NIST_QC2_', '13_CTRL_1_']		# naming pattern of replicates
	elif mousemodel==4:	# STAM
		# STAM:
		repnumvolcano=6
		negdataname=['1_CIT_1_', '2_CIT_2_', '3_CIT_3_', '4_CIT_4_', '5_CIT_5_', '6_CIT_6_', 
				'7_STAH_1_', '8_STAH_2_', '9_STAH_3_', '10_STAH_4_', '11_STAH_5_', '12_STAH_6_', 
				'14_CTRL_2_', '15_CTRL_3_', '16_CTRL_4_', '17_CTRL_5_', '18_CTRL_6_', 
				'20_CDDA_2_', '21_CDDA_3_', '22_CDDA_4_', '23_CDDA_5_', '24_CDDA_6_', 
				'24b_CDDA_6_', '24c_CDDA_6_', 
				'NIST_QC1_', 'NIST_QC2_', '13_CTRL_1_']		# naming pattern of replicates
	elif mousemodel==5:
		negdataname=['NIST1_', 'NIST2_', 'NIST3_']
	elif mousemodel==6:
		negdataname=['JQC_', 'JQC2_', 'JQC3_']
	elif mousemodel==7:			# liver interday validation
		negdataname=['KML_', 'KML24d_', 'KML24e_']
	elif mousemodel==8:	# MC4R-KO_HFD
		repnumvolcano=3
		negdataname=['FB1_', 'FB2_', 'FB3_', 'NIST_QC1_']		# naming pattern of replicates
	else:
		print('Edit/add sampleset and naming pattern with additional elif clause!')
		quit()

	################################################# MODIFY 


	# ms1splitlist contains those species, for which double bond or branching isomers are chromatographically separated with no interference from multiple isomer species and with 
	# low abundance overall (missing or inconsistent MS2 data), so that isomer splitting of quantities according to isomers is based on MS1 precursor integrals of the peaks
	if mousemodel==2:		# MC4R-KO
		ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_33:2_(15:0_18:2)', 'PC_35:1_(17:0_18:1)', 
					'PC_35:2_(17:0_18:2)', 'PC_36:4_(16:0_20:4)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 
					'PC_38:7_(16:1_22:6)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)', 
					'PC_36:3_(16:0_20:3)', 'PC_38:3_(18:0_20:3)']
	elif mousemodel>2:		# CDDA and STAM
		if mousemodel==5:
			ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_32:1_(16:0_16:1)', 'PC_33:0_(15:0_18:0)', 'PC_33:0_(16:0_17:0)', 
						'PC_33:2_(15:0_18:2)', 'PC_35:1_(17:0_18:1)', 'PC_35:2_(17:0_18:2)', 
						'PC_36:3_(16:0_20:3)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 'PC_38:3_(18:0_20:3)', 
						'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)']
		elif mousemodel==6:
			ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_32:1_(16:0_16:1)', 'PC_33:0_(15:0_18:0)', 'PC_33:0_(16:0_17:0)', 
						'PC_33:2_(15:0_18:2)', 'PC_35:1_(17:0_18:1)', 'PC_35:2_(17:0_18:2)', 
						'PC_36:3_(16:0_20:3)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 'PC_38:3_(18:0_20:3)', 
						'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)']
		elif mousemodel==7:
			ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_32:1_(16:0_16:1)', 'PC_33:0_(15:0_18:0)', 'PC_33:0_(16:0_17:0)', 
						'PC_33:2_(15:0_18:2)', 'PC_35:1_(17:0_18:1)', 'PC_35:2_(17:0_18:2)', 
						'PC_36:3_(16:0_20:3)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 'PC_38:3_(18:0_20:3)', 
						'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)']
		elif mousemodel==8:
			ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_33:2_(15:0_18:2)', 'PC_35:1_(17:0_18:1)', 
					'PC_35:2_(17:0_18:2)', 'PC_36:4_(16:0_20:4)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 
					'PC_38:7_(16:1_22:6)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)', 
					'PC_36:3_(16:0_20:3)', 'PC_38:3_(18:0_20:3)']
		else:
			ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_32:1_(16:0_16:1)', 'PC_33:0_(15:0_18:0)', 'PC_33:0_(16:0_17:0)', 
						'PC_33:2_(15:0_18:2)', 'PC_35:1_(17:0_18:1)', 'PC_35:2_(17:0_18:2)', 
						'PC_36:3_(16:0_20:3)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 'PC_38:3_(18:0_20:3)', 
						'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)']
	elif mousemodel==1:		# CCl4
		ms1splitlist=['PC_31:0_(15:0_16:0)', 'PC_33:2_(15:0_18:2)', 'PC_34:3_(16:0_18:3)', 'PC_35:1_(17:0_18:1)', 
					'PC_35:2_(17:0_18:2)', 'PC_36:4_(16:0_20:4)', 'PC_37:4_(17:0_20:4)', 'PC_37:2_(19:0_18:2)', 
					'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:5_(18:0_22:5)', 'PE_36:3_(16:0_20:3)', 
					'PC_36:3_(16:0_20:3)', 'PC_38:3_(18:0_20:3)', 'PE_38:3_(18:0_20:3)']


	# remergelist contains those species that have faulty quantities based on SKYLITE_3 analysis, because of missing MS2 data for some isomer of the sum composition 
	# In this case the original total MS1 integral is summed up again for the sum composition and then an MS1 based splitting is applied to generate the quantities of the 
	# (chromatographically resolved) isomers
	if mousemodel==2:
		remergelist=['PC_38:7_(18:2_20:5)', 'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)']
	elif mousemodel>2:
		if mousemodel==5:
			remergelist=[]
		elif mousemodel==6:
			remergelist=['PC_38:7_(18:2_20:5)', 'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:7_(20:3_20:4)', 'PC_40:7_(18:1_22:6)']
		elif mousemodel==7:
			remergelist=['PC_38:7_(18:2_20:5)', 'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)',  
				'OC_38:4_(16:0_22:4)', 'OC_38:4_(18:0_20:4)', 'QC_36:3_(18:1_18:2)', 'QC_36:3_(18:2_18:1)', 'PE_36:2_(18:0_18:2)', 'PE_36:2_(18:1_18:1)', 
				'PE_36:3_(16:0_20:3)', 'PE_36:3_(18:1_18:2)', 'QE_38:5_(16:1_22:4)', 'QE_38:5_(18:1_20:4)', 'QE_40:5_(18:1_22:4)', 'QE_40:5_(20:1_20:4)',
				'QE_40:6_(18:1_22:5)', 'QE_40:6_(18:2_22:4)']
		elif mousemodel==8:
			remergelist=['PC_38:7_(18:2_20:5)', 'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)']
		else:
			remergelist=['PC_38:2_(18:0_20:2)', 'PC_38:2_(20:0_18:2)', 'PC_38:2_(18:1_20:1)', 
					'PC_38:7_(18:2_20:5)', 'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:4_(18:0_22:4)', 
					'PC_40:4_(20:0_20:4)','PC_40:7_(20:3_20:4)', 'PC_40:7_(18:1_22:6)', 'PE_36:2_(18:0_18:2)', 
					'PE_36:2_(18:1_18:1)', 'PE_36:4_(16:0_20:4)', 'PE_36:4_(18:2_18:2)']
	elif mousemodel==1:
		remergelist=['PC_38:7_(18:2_20:5)', 'PC_38:7_(16:1_22:6)', 'PC_38:7_(16:0_22:7)', 'PC_40:7_(20:3_20:4)', 'PC_40:7_(18:1_22:6)']


	# begin read Skyline reports file pos
	#trdf=pd.read_csv('Skyline_Report_JPM_ILS_quantification_pos.csv')
	#trdf=pd.read_csv('Skyl_Rep_JPM_NIST_USP_pos_curated.csv')
	#toprowp=[trdf.columns.values.tolist()]
	#toprowp=toprowp[0]
	#trdf=trdf.transpose()
	#posreport=trdf.values.tolist()
	#print('Number of rows in Skyline_Report_JPM_ILS_quantification_pos.csv: %d' % ki)
	# begin read Skyline reports file neg
	ntrdf=pd.read_csv('Skyline_Report_isomers_for_SKYLITE_4_neg.csv', low_memory=False)		# SKYLINE REPORT ISOMER RESOLVED IS NOW INPUT FILE
	toprown=[ntrdf.columns.values.tolist()]
	toprown=toprown[0]
	ntrdf=ntrdf.transpose()
	negreport=ntrdf.values.tolist()
	#print('Number of rows in Skyline_Report_JPM_ILS_quantification_neg.csv: %d' % nki)
	# end read  Skyline reports files

	outfilename='SKYLITE_3_quantities_for_SKYLITE_4_neg.xlsx'	# OUTPUT FILE FROM PREVIOUS STEP IS NOW INPUT FILE
	wb=openpyxl.load_workbook(outfilename)
	ws=wb['Quantification']
	ws2=wb.create_sheet('Quantification_isomers')


############################################################################################################################
############################################################################################################################
investigatelipid='OC_38:4_(16:0_22:4)'
investigatereplicate='KML24d_'
############################################################################################################################
############################################################################################################################
posdataname=negdataname #['NIST1', 'NIST2', 'NIST3']
goff=0					################################# GENERAL OFF SWITCH FOR ALL EXCEPT BLANK ### TURN TO 0 FOR USE OF ALL
chk=1
testrun=0 # default 0, as replicate information needs to be loaded	# IF MANUAL USAGE FOR LOI: SET testrun=1
checkup=0
after=str(beforeall)
tnow=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+' / '+after[11]+after[12]+after[13]+after[14]+after[15]+after[16]+after[17]+after[18]
print('Calculation start at %s' % tnow)


################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
iabh=0.999885
iabc=0.9893
iabn=0.996
iabo=0.99636
iabp=1
###############################################################################################################################


# go through negreport to assign, which entry / index in list belongs to an isomer that is isomer 1, 2, 3...
ixlist=[]			# ixlist indicates if a lipid is only 1 isomer or isomer 1 or 2 or 3, ixlist[ix] is parallel to negreport[0][ix]
ix=1
ixlist.append(ix)
ixi=1
while ixi<len(negreport[0]):
	if negreport[1][ixi]==negreport[1][ixi-1]:
		if negreport[6][ixi]==negreport[6][ixi-1]:
			ix=ix
		elif str(negreport[6][ixi])=='precursor':
			ix=ix+1
		else:
			ix=ix
	else:
		ix=1
	ixlist.append(ix)
	#if ix>0:
	#	if '20_CDDA_2_' in str(negreport[20][ixi]):
	#		print(str(negreport[1][ixi])+'_'+str(ix))
	ixi=ixi+1 
#quit()

# search for lipid that requires splitting of existing quantity over chromatographically resolved isomers

#go through previous report and write toprow for new sheet, 
# then go through lipids and apply splitting where required (splitting of quantities done based on MS2 integrals, exceptions noted in ms1splitlist)
# lipids in ms1splitlist will be quantified based on the isomer-associated ms1 integral

ws2q3s4.cell(row=1, column=1).value='Lipid'
ws2q3s4.cell(row=1, column=2).value='RT [min]'
ccol=3
while (ccol-3)<len(negdataname):
	if tissuetype==1:
		repname='c_'+str(negdataname[ccol-3])+' [nmol/mg]' #'replicate '+str(ccol-1)
	else:
		repname='c_'+str(negdataname[ccol-3])+' [nmol/mL]' #'replicate '+str(ccol-1)
	ws2q3s4.cell(row=1, column=ccol).value=repname
	ccol=ccol+1

t=2	# index to read previous excel sheet
w=2	# index to write in new excel sheet
#lsl=0
go=1
while go==1:
	#lsl=lsl+1
	clipid=ws1q3s4.cell(row=t, column=1).value
	clipid=str(clipid)
	slipid=clipid
	isocount=0
	isorlist=[]
	isortlist=[]
	r=0
	while r<len(negreport[0]):
		if 'PC-O' in str(clipid):
			clipid=clipid[4:]
			clipid='OC'+clipid
		elif 'PC-O/P' in str(clipid):
			clipid=clipid[6:]
			clipid='QC'+clipid
		elif 'PE-O' in str(clipid):
			clipid=clipid[4:]
			clipid='OE'+clipid
		elif 'PE-O/P' in str(clipid):
			clipid=clipid[6:]
			clipid='QE'+clipid
		elif 'LPC' in str(clipid):
			clipid=clipid[3:]
			clipid='LC'+clipid
		elif 'LPE' in str(clipid):
			clipid=clipid[3:]
			clipid='LE'+clipid	
		elif 'LPI' in str(clipid):
			clipid=clipid[3:]
			clipid='LI'+clipid
		if str(clipid)==negreport[1][r]:	#			
			if str(negreport[6][r])=='precursor':
				if negdataname[0] in str(negreport[20][r]):
					isorlist.append(r)			# list with indeces r in negreport of precursor of rep 1 for each isomer
					isortlist.append(str(negreport[12][r]))
					crt=str(negreport[12][r])
					if isocount>0:
						ok=1
						#print(clipid)
						#print(slipid)
						#print('#############')
					isocount=isocount+1
					# get MS2 quantity to add to cms2sumlist for splitting if required
					
		r=r+1
	if isocount>1:				# multiple isomers found (e.g., multiple PC(16:0_20:3) species found)
		repi=0
		while repi<len(negdataname):
			
			# check, if species is in ms1splitlist, then split by MS1, otherwise split by MS2
			if clipid in ms1splitlist:
				if clipid==investigatelipid:
					if negdataname[repi]==investigatereplicate:
						print('---')
						print(clipid)
						print(negdataname[repi])
						print('# Isomers found: %d' % len(isorlist))

				#if clipid in remergelist:
				#	# re-merge before splitting by MS1
	 			#	ok=1 


				# split by MS1																			### CHECK steps
				# get MS1 integrals of this lipid
				#print(str(negdataname[repi]))
				cms1sumlist=[]
				z=0
				while z<len(isorlist):
					r=isorlist[z]
					crt=isortlist[z]
					cms1sum=0
					cist=negreport[18][r]	# current isomer start time
					ciet=negreport[19][r]	# current isomer end time
					#set s to make sure to catch correct precursor
					s=r
					sgo=1
					while sgo==1:
						if 'precursor' in str(negreport[6][s]):
							sgo=1
							s=s-1
						else:
							sgo=0
					if clipid==investigatelipid:
						if negdataname[repi]==investigatereplicate:
							print('###')
							print('r = %d' % r)
							print('RT Start = %s' % cist)
							print('RT End = %s' % ciet)
					gofa=1
					while gofa==1:
						if 'precursor' in str(negreport[6][s]):
							if str(negdataname[repi]) in str(negreport[20][s]):
								if clipid==investigatelipid:
									if negdataname[repi]==investigatereplicate:
										print('...')
										print('s = %d' % s)
										#print(negreport[20][s])
										print('Precursor integral = %.2f ' % negreport[13][s])
								cms1sum=cms1sum+float(negreport[13][s])
						if abs(cist-negreport[18][s])<0.04:
							if abs(ciet-negreport[19][s])<0.04:
								gofa=gofa
							else:
								gofa=1
						else:
							gofa=1
						if str(negreport[6][s])=='precursor':
							if str(negdataname[repi]) in str(negreport[20][s]):
								gofa=0
						s=s+1
					cms1sumlist.append(cms1sum)
					z=z+1
				# split and write results
				if clipid==investigatelipid:
					if negdataname[repi]==investigatereplicate:
						print(cms1sumlist)
				#print(cms2sumlist)
				z=0
				while z<len(isorlist):
					cquant=ws1q3s4.cell(row=t, column=(repi+3)).value
					wlipid=slipid+'_isomer_'+str(z+1)
					ws2q3s4.cell(row=w+z, column=1).value=wlipid
					crt=isortlist[z]
					ws2q3s4.cell(row=w+z, column=2).value=crt			# RT
					if sum(cms1sumlist)==0:
						wquant=cquant
					elif cquant is None:
						wquant=cquant
					else:
						#print(cquant)
						#print(cms1sumlist[z])
						#print((sum(cms1sumlist)))
						wquant=cquant*(cms1sumlist[z]/sum(cms1sumlist))
					if wquant==0:
						ok=1
					else:
						ws2q3s4.cell(row=w+z, column=(repi+3)).value=wquant
					z=z+1
				repi=repi+1

			else:
				# get MS2 integrals of all isomer features of this lipid
				#print(str(negdataname[repi]))
				cms2sumlist=[]
				z=0
				while z<len(isorlist):
					r=isorlist[z]
					crt=isortlist[z]
					cms2sum=0

					# new module based on ixlist for assignment of isomer number
					grb=0
					while grb<len(negreport[1]):
						if str(negdataname[repi]) in str(negreport[20][grb]):	# replicate name
							if negreport[1][grb]==clipid:					# lipid name
								if (z+1)==ixlist[grb]:						# isomer number
									if 'FA' in str(negreport[6][grb]):		# FA diagnostic fragment
										cms2sum=cms2sum+float(negreport[13][grb])
										if clipid==investigatelipid:
											if negdataname[repi]==investigatereplicate:
												prc=0
												if prc==1:
													print(',,,,')
													#print(repi)
													print(negdataname[repi])
													print(clipid)
													print(z)
													print(float(negreport[13][grb]))
													print(grb)
													print(ixlist[grb])
						grb=grb+1
					
					cms2sumlist.append(cms2sum)
					z=z+1
				# split and write results
				#print(cms2sumlist)
				z=0
				while z<len(isorlist):
					cquant=ws1q3s4.cell(row=t, column=(repi+3)).value
					wlipid=slipid+'_isomer_'+str(z+1)
					ws2q3s4.cell(row=w+z, column=1).value=wlipid
					crt=isortlist[z]
					ws2q3s4.cell(row=w+z, column=2).value=crt			# RT
					if sum(cms2sumlist)==0:
						wquant=cquant
					elif cquant is None:
						wquant=cquant
					else:
						#print(cquant)
						#print(cms2sumlist[z])
						#print((sum(cms2sumlist)))
						wquant=cquant*(cms2sumlist[z]/sum(cms2sumlist))
					if wquant==0:
						ok=1
					else:
						ws2q3s4.cell(row=w+z, column=(repi+3)).value=wquant
					if clipid==investigatelipid:
						prc=0
						if prc==1:
							if negdataname[repi]==investigatereplicate:
								print('###')
								print('wquant = %.6f' % wquant)
								print(cms2sumlist)
								print(z)
							else:
								print('###')
								print('wquant = %.6f' % wquant)
								print(cms2sumlist)
								print(z)
					z=z+1
				repi=repi+1
				
		w=w+len(isorlist)
	
	else:  
		ws2q3s4.cell(row=w, column=1).value=slipid	# copy lipid name and RT without any splitting or modification
		ws2q3s4.cell(row=w, column=2).value=crt
		repi=0
		while repi<len(negdataname):
			# check for remergelist to correct splitting ???
			# lipid naming check for etherlipids
			if clipid==investigatelipid:
				if negdataname[repi]==investigatereplicate:
					print(clipid)


			if clipid in remergelist:
				#
				#print('....')
				#print(clipid)
				#print(negdataname[repi])
				#print(repi)
				# retrieve original sum of quantities for the current sum composition lipid (e.g., abs quantity of sum of all PC 38:7 isomers)
				# go through excel file
				originalsum=0
				ccl=repi+3
				crw=2
				cscs=clipid[0]+clipid[1]+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]	# current sum composition
				if clipid[0]+clipid[1]=='OC':
					csc='PC-O'+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]	# current sum composition
				elif clipid[0]+clipid[1]=='QC':
					csc='PC-O/P'+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]	# current sum composition
				elif clipid[0]+clipid[1]=='OE':
					csc='PE-O'+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]	# current sum composition
				elif clipid[0]+clipid[1]=='QE':
					csc='PE-O/P'+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]	# current sum composition
				else:
					csc=clipid[0]+clipid[1]+clipid[2]+clipid[3]+clipid[4]+clipid[5]+clipid[6]	# current sum composition

				gorw=1
				while gorw==1:
					selipid=ws1q3s4.cell(row=crw, column=1).value
					if selipid is None:
						gorw=0
					elif str(selipid)=='':
						gorw=0
					else:
						if csc in selipid:
							oquant=ws1q3s4.cell(row=crw, column=ccl).value
							if oquant is None:
								oquant=0
							originalsum=originalsum+oquant
					crw=crw+1
				#print(originalsum)
				#quit()
				# get MS1 precursor integral of all lipid isomers within sum composition as well as current lipid
				cms1list=[]
				cms1i=0
				sr=0
				#isn=1
				while sr<len(negreport[0]):
					if cscs in str(negreport[1][sr]):
						if 'precursor'==str(negreport[6][sr]):
							if str(negdataname[repi]) in str(negreport[20][sr]):
								cms1list.append(negreport[13][sr])
								if clipid==negreport[1][sr]:
									cms1i=negreport[13][sr]
									#isn=isn+1
					sr=sr+1
				# calc fraction of MS1 precursor intensities and apply to original sum composition quantity, then write in file
				nq=cms1i/sum(cms1list)*originalsum
				if clipid==investigatelipid:
					if negdataname[repi]==investigatereplicate:
						print(originalsum)
						print(slipid)
				#print(cms1i)
				#print(cms1list)
				#print(nq)
				#quit()
				if nq==0:
					ok=1
				else:
					ws2q3s4.cell(row=w, column=repi+3).value=nq

			else:
				cquant=ws1q3s4.cell(row=t, column=(repi+3)).value
				if cquant==0:
					ok=1
				else:
					ws2q3s4.cell(row=w, column=(repi+3)).value=cquant
			repi=repi+1
		w=w+1
	t=t+1
	clipid=ws1q3s4.cell(row=t, column=1).value
	if clipid is None:
		go=0
	elif str(clipid)=='NaN':
		go=0
	elif str(clipid)=='':
		go=0
	else:
		go=1




after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'

runidentifier=ws3.cell(row=2, column=3).value
runidentifier=str(runidentifier)
if today in runidentifier:
	outfilename=runidentifier+'SKYLITE_Q4_neg.xlsx'
else:
	outfilename=today+runidentifier+'SKYLITE_Q4_neg.xlsx'


#outfilename=outfilename[:-10]
#outfilename=today+'SKYLITE_4_isomer_quantities_neg.xlsx'
wbq3s4.save(outfilename)
#################################################################################################################################################################
#################################################################################################################################################################
if excelinput==0:
	if mousemodel==7:
		tissuetype==0
		repnumvolcano=0

if tissuetype==1:
	#################################################################################################################################################################
	#################################################################################################################################################################
	wbq3s4=openpyxl.load_workbook(outfilename)
	#ws2q3s4=wbq3s4.create_sheet('Quantification_isomers')
	ws2q3s4=wbq3s4['Quantification_isomers']

	lplist=[]	# lipid list
	pvlist=[]	# raw p value
	ablist=[]	# mean abundance of lipid in set of replicates with higher abundance
	lgfclist=[]	# log2 fold change list
	# Begin calculate means, standard deviations, fold changes and p-values
	snum=repnumvolcano #6	# sample number for statistics calculation (e.g., comparison of 6 WT with 6 KO samples, these have to be the first 6 (12) samples in list negdataname)
	csnum=snum
	go=1
	r=2
	while go==1:
		csnum=snum
		clipid=ws2q3s4.cell(row=r, column=1).value
		asum=0
		bsum=0
		alist=[]
		blist=[]
		sni=0
		while sni<snum:
			caq=ws2q3s4.cell(row=r, column=sni+3).value
			if caq is None:
				caq=0
				csnum=csnum-1
			else:
				cbq=ws2q3s4.cell(row=r, column=sni+3+snum).value
				if cbq is None:
					cbq=0
					csnum=csnum-1
				else:
					caq=float(caq)
					asum=asum+caq
					alist.append(caq)
					cbq=float(cbq)
					bsum=bsum+cbq
					blist.append(cbq)
			sni=sni+1
		if csnum==0:
			amean=0
			bmean=0
		else:
			amean=asum/csnum
			bmean=bsum/csnum
		#print('...')
		#print(clipid)
		#print(alist)
		#print(blist)

		# begin shorten clipid name to remove sum composition from name of PC, PE and PI, shorten isomer to i 
		clipid=str(clipid)
		if 'isomer' in clipid:
			ii=clipid[len(clipid)-1]
			clipid=clipid[:-7]+ii
		shrt=0
		if 'PC' in clipid:
			shrt=1
		elif 'PE' in clipid:
			shrt=1
		elif 'PI' in clipid:
			shrt=1
		if 'L' in clipid:
			shrt=0
		if shrt==1:
			ni=0
			while ni<len(clipid):
				if clipid[ni]==':':
					clipid=clipid[:-(len(clipid)-ni+3)]+clipid[-((len(clipid)-ni)-3):]
					ni=len(clipid)
				ni=ni+1
		# begin shorten clipid name to remove sum composition from name of PC, PE and PI, shorten isomer to i 

		ws2q3s4.cell(row=r, column=3+len(negdataname)+2).value=clipid
		ws2q3s4.cell(row=r, column=3+len(negdataname)+3).value=amean    #write 
		ws2q3s4.cell(row=r, column=3+len(negdataname)+5).value=bmean    #write 
		if len(alist)<2:
			astd=''
		else:
			astd=statistics.stdev(alist)
		if len(blist)<2:
			bstd=''
		else:
			bstd=statistics.stdev(blist)
		ws2q3s4.cell(row=r, column=3+len(negdataname)+4).value=astd    #write 
		ws2q3s4.cell(row=r, column=3+len(negdataname)+6).value=bstd    #write 
		if amean==0:
			fcm=0
		else:
			fcm=bmean/amean
			ws2q3s4.cell(row=r, column=3+len(negdataname)+7).value=fcm    #write 
		#print(fcm)
		if fcm==0:
			lgfcm=0
		else:
			lgfcm=math.log2(fcm)
			ws2q3s4.cell(row=r, column=3+len(negdataname)+9).value=lgfcm    #write 
		lgfclist.append(lgfcm)
		if lgfcm==0:
			ok=1
			# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
			lplist.append(clipid)
			ablist.append(0)
			pvlist.append(1)
			# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
		else:
			# begin calc P value
			t,p=stats.ttest_ind(alist, blist, equal_var=False)
			ts=str(t)
			ps=str(p)
			if ts=='nan':
				t=str(1.0)
			if ps=='nan':
				p=str(1.0)
			t=float(t)
			p=float(p)
			#print(p)
			#print(t)
			# write P value and ttest statistic in excel output file
			ws2q3s4.cell(row=r, column=3+len(negdataname)+8).value=p    #write 
			ws2q3s4.cell(row=r, column=3+len(negdataname)+12).value=t    #write 
			clgp=math.log10(p)	# p value for volcano plot as -log10()
			clgp=(-1)*clgp
			sign=0
			if p<0.05:		# determine if fold change and p value significant
				if fcm>2:
					sign=1
				elif fcm<0.5:
					sign=1
			if sign==1:
				ws2q3s4.cell(row=r, column=3+len(negdataname)+11).value=clgp    #write significant p values / with significant fold changes
			else:
				if clgp==0:
					ok=1
				else:
					ws2q3s4.cell(row=r, column=3+len(negdataname)+10).value=clgp    #write insignificant
			# end calc P value
			# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
			lplist.append(clipid)
			if amean>bmean:
				lmean=amean
			else:
				lmean=bmean
			ablist.append(lmean)
			pvlist.append(p)
			# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down

		r=r+1
		clipid=ws2.cell(row=r, column=1).value
		if clipid is None:
			go=0
		elif str(clipid)=='NaN':
			go=0
		elif str(clipid)=='':
			go=0
		else:
			go=1


	ws2q3s4.cell(row=1, column=3+len(negdataname)+2).value='Lipid'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+3).value='Mean [col C-H]'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+4).value='Std. dev. [col C-H]'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+5).value='Mean [col I-N]'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+6).value='Std. dev. [col I-N]'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+7).value='Fold change of mean'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+8).value='P value'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+9).value='log2 fold change'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+10).value='="-log10 p-value (raw, not significant)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+11).value='="-log10 p-value (raw, significant)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+12).value='t-test statistic'

	# End calculate means, standard deviations, fold changes and p-values

	# Begin calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down
	ws2q3s4.cell(row=1, column=3+len(negdataname)+13).value='P value (Bonferroni)'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+14).value='P value (BH)'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+15).value='P value (ASD)'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+16).value='="-log10 p-value (Bonferroni)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+17).value='="-log10 p-value (ASD)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+18).value='="-log10 p-value (BH; insignificant)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+19).value='="-log10 p-value (BH; also significant after ASD and BF)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+20).value='="-log10 p-value (BH; also significant after ASD, but not BF)"'
	ws2q3s4.cell(row=1, column=3+len(negdataname)+21).value='="-log10 p-value (BH; significant, but not after ASD or BF)"'
	nmc=len(lplist)


	#print('Lists for BH:')
	# begin create sorted lists by p value for BH
	#print(lplist)
	#print(pvlist)
	psortlplist=[]
	psortpvlist=[]
	dlplist=[]
	dpvlist=[]
	ni=0
	while ni<len(lplist):
		dlplist.append(lplist[ni])
		dpvlist.append(pvlist[ni])
		ni=ni+1
	while len(dlplist)>0:
		cpv=min(dpvlist)
		cdi=dpvlist.index(cpv)
		psortpvlist.append(dpvlist[cdi])
		psortlplist.append(dlplist[cdi])
		del dpvlist[cdi]
		del dlplist[cdi]
	#print(lplist)
	#print(pvlist)
	#print(psortlplist)
	#print(psortpvlist)
	# end create sorted lists by p value for BH
	#print('Lists for ASD:')
	# begin create sorted lists by p value for ASD
	#print(lplist)
	#print(pvlist)
	#print(ablist)
	absortlplist=[]
	absortpvlist=[]
	absortablist=[]
	dlplist=[]
	dpvlist=[]
	dablist=[]
	ni=0
	while ni<len(lplist):
		dlplist.append(lplist[ni])
		dpvlist.append(pvlist[ni])
		dablist.append(ablist[ni])
		ni=ni+1
	while len(dlplist)>0:
		cab=min(dablist)
		cdi=dablist.index(cab)
		absortpvlist.append(dpvlist[cdi])
		absortablist.append(dablist[cdi])
		absortlplist.append(dlplist[cdi])
		del dpvlist[cdi]
		del dlplist[cdi]
		del dablist[cdi]
	#print(lplist)
	#print(pvlist)
	#print(absortlplist)
	#print(absortpvlist)
	#print(absortablist)
	# end create sorted lists by p value for ASD
	apvi=0
	while apvi<len(lplist):
		if pvlist[apvi]==1:
			ok=1
		elif pvlist[apvi] is None:
			ok=1
		else:
			# begin Bonferroni
			bfpv=pvlist[apvi]*nmc
			if bfpv>1:
				bfpv=1
			logbfpv=math.log10(bfpv)	# p value for volcano plot as -log10()
			logbfpv=(-1)*logbfpv
			ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+13).value=bfpv
			ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+16).value=logbfpv
			# end Bonferroni

			# begin Abundance-Step-Down
			kasd=(absortlplist.index(lplist[apvi]))
			asdpv=pvlist[apvi]*(nmc-kasd)
			if asdpv>1:
				asdpv=1
			logasdpv=math.log10(asdpv)	# p value for volcano plot as -log10()
			logasdpv=(-1)*logasdpv
			ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+14).value=asdpv
			ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+17).value=logasdpv
			# end Abundance-Step-Down

			# begin Benjamini-Hochberg
			kbh=(psortlplist.index(lplist[apvi]))+1
			bhpv=pvlist[apvi]*(nmc/kbh)
			if bhpv>1:
				bhpv=1
			logbhpv=math.log10(bhpv)	# p value for volcano plot as -log10()
			logbhpv=(-1)*logbhpv
			ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+15).value=bhpv
			# determine colour code (column), depending on: 
			# 1) Insignificant after BH (18)
			# 2) Significant after BH and significant after BF (19)
			# 3) Significant after BH and significant after ASD, but not after BF (20)
			# 4) Significant after BH, but not after ASD or BF (21)
			if abs(lgfclist[apvi])<1:
				ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+18).value=logbhpv	# 1) Insignificant due to small fold change
			else:
				if bhpv>0.05:
					ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+18).value=logbhpv	# 1) Insignificant after BH
				else:
					if bfpv<0.05:
						ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+19).value=logbhpv	# 2) Significant after BH and significant after BF and ASD
					else:
						if asdpv<0.05:
							ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+20).value=logbhpv	# 3) Significant after BH and significant after ASD, but not after BF
						else:
							ws2q3s4.cell(row=2+apvi, column=3+len(negdataname)+21).value=logbhpv	# 4) Significant after BH, but not after ASD or BF 
			# end Benjamini-Hochberg

		apvi=apvi+1
	# End calculate adjusted P values according to Bonferroni, Benjamini-Hochberg and Abundance-Step-Down


	runidentifier=ws3.cell(row=2, column=3).value
	runidentifier=str(runidentifier)
	if today in runidentifier:
		outfilename=runidentifier+'SKYLITE_Q4_neg.xlsx'
	else:
		outfilename=today+runidentifier+'SKYLITE_Q4_neg.xlsx'
	wbq3s4.save(outfilename)

	print('Calculation completed. The output file with isomer resolved quantities is saved as: %s' % outfilename)
	



